from app.database.models import async_session
import app.database.models as tb
from sqlalchemy import select, and_, update, func
from sqlalchemy.orm import joinedload, selectinload
import logging
from datetime import datetime, timedelta
from aiogram.utils.markdown import hbold, hunderline
from aiogram.types import BufferedInputFile
from aiogram import Bot
from typing import List
import pandas as pd
from io import BytesIO
from openpyxl.utils import get_column_letter
import app.keyboards as kb
from sqlalchemy.ext.asyncio import AsyncSession
from openpyxl.styles import Alignment
import matplotlib.pyplot as plt
from cachetools import TTLCache

from staticmap import StaticMap, Line
import asyncio

user_cache = TTLCache(maxsize=100, ttl=300)

def connection(func):
    async def inner(*args, **kwargs):
        async with async_session() as session:
            try:
                result = await func(session, *args, **kwargs)
                await session.commit()
                return result
            except Exception as e:
                await session.rollback()
                logging.error(f"ÐžÑˆÐ¸Ð±ÐºÐ° Ð² Ñ„ÑƒÐ½ÐºÑ†Ð¸Ð¸ {func.__name__}: {e}")
                raise e
            finally:
                await session.close()
    return inner

@connection
async def check_user(session: AsyncSession, tg_id)-> bool:
    user = await session.scalar(select(tb.User).where(tb.User.tgId == tg_id))

    return user is not None


@connection
async def reg_user(session: AsyncSession, data, tg_id)-> None:
    role = await session.scalar(select(tb.Role).where(tb.Role.roleName == data['role']))

    new_user = tb.User(
        tgId = tg_id,
        phone=data.get("number"),
        fio=data.get("fio"),
        roleId = role.idRole
    )

    session.add(new_user)

@connection
async def get_cargo_types(session: AsyncSession):
    result = await session.execute(select(tb.CargoType).order_by(tb.CargoType.cargoTypeName))
    cargo_types = result.scalars().all()
    return {cargo.idCargoType: cargo.cargoTypeName for cargo in cargo_types}

@connection
async def get_cargo_type_name_by_id(session: AsyncSession, data):
    cargo_type_name = await session.scalar(select(tb.CargoType).where(tb.CargoType.idCargoType == data))

    return cargo_type_name.cargoTypeName



@connection
async def add_new_order(session: AsyncSession, data):
    disp_id = await session.scalar(select(tb.User).where(tb.User.tgId == data["tg_id"]))
    new_order = tb.Order(
        cargoName=data["cargo_name"],
        cargoDescription=data["cargo_description"],
        cargoTypeId=int(data["cargo_type_id"]),
        cargo_weight=float(data["cargo_weight"]),
        depart_loc=data["depart_loc"],
        goal_loc=data["goal_loc"],
        time=datetime.strptime(data["time"], '%H:%M %d.%m.%Y'),
        orderStatusId = 1,
        dispatcherId = disp_id.idUser,
        driverId=None,
        create_order_time = datetime.now(),
    )
    if "photoId" in data:
        new_order.photoId = data["photoId"]
    if "isUrgent" in data:
        new_order.isUrgent = data["isUrgent"]

    session.add(new_order)
    await session.flush()
    await session.refresh(new_order)
    return new_order.idOrder

@connection
async def alarm_for_drivers(session: AsyncSession, orderId, bot: Bot):
    drivers = await session.scalars(select(tb.User).where(tb.User.roleId == 2))
    order = await session.scalar(select(tb.Order).options(joinedload(tb.Order.cargoType)).where(tb.Order.idOrder == orderId))
    mes = f'Ð¡Ñ€Ð¾Ñ‡Ð½Ñ‹Ð¹ Ð·Ð°ÐºÐ°Ð·:\n\n' + await form_order(order=order, cargo_type=order.cargoType.cargoTypeName)
    for driver in drivers:
        await bot.send_message(driver.tgId, mes, reply_markup=await kb.alarm_kb(orderId=orderId), parse_mode="HTML")

statuses = {
    1: "Ð”Ð¾ÑÑ‚ÑƒÐ¿ÐµÐ½",
    2: "Ð’ Ñ€Ð°Ð±Ð¾Ñ‚Ðµ",
    3: "Ð—Ð°Ð²ÐµÑ€ÑˆÐµÐ½",
    4: "ÐžÑ‚Ð¼ÐµÐ½ÐµÐ½"
}

@connection
async def get_order_keys(session: AsyncSession, dateTime: datetime = None, tg_id = None, isActual = False, isPrivateCatalog =False, statusId:int = None):
    stmt = select(tb.Order)
    user = await session.scalar(select(tb.User).where(tb.User.tgId == tg_id))
    if not isPrivateCatalog:
        if(dateTime != None):
            start_time = dateTime
            end_time = dateTime + timedelta(days=1)
            stmt = stmt.where(and_(tb.Order.time > start_time,
                                    tb.Order.time < end_time))
        if user.roleId == 2:
            stmt = stmt.where(tb.Order.orderStatusId == 1)
        stmt = stmt.order_by(tb.Order.time)

    else:
        match user.roleId:
            case 1: #Ð”Ð¸ÑÐ¿ÐµÑ‚Ñ‡ÐµÑ€
                role_condition= tb.Order.dispatcherId == user.idUser
                
            case 2:  # Ð’Ð¾Ð´Ð¸Ñ‚ÐµÐ»ÑŒ
                role_condition = tb.Order.driverId == user.idUser
            case 3:
                role_condition= tb.Order.dispatcherId == user.idUser
            case _:
                raise ValueError(f"Ð Ð¾Ð»ÑŒ {user.roleId} Ð½Ðµ Ð¿Ð¾Ð´Ð´ÐµÑ€Ð¶Ð¸Ð²Ð°ÐµÑ‚ÑÑ")
        
        stmt = stmt.where(role_condition)

        if isActual:
            if statusId in [1,2]:
                stmt = stmt.where(tb.Order.orderStatusId == statusId)
            else:
                stmt = stmt.where(tb.Order.orderStatusId.in_([1,2]))
            statusId = None
        
        stmt = stmt.order_by(tb.Order.time.desc())
    
    if statusId is not None:

        stmt = stmt.where(tb.Order.orderStatusId == statusId)    

    result = await session.execute(stmt)
    orders = result.scalars().all()    

    
    order_keys = []
    for order in orders:
        order_keys.append(order.idOrder)
    return order_keys

@connection
async def get_orders(session: AsyncSession, ordersKeys, start: int, end: int):
    actiual_order_list = ordersKeys[start:end]
    stmt = (
        select(tb.Order)
        .options(joinedload(tb.Order.executor))  # Ð—Ð°Ð³Ñ€ÑƒÐ¶Ð°ÐµÐ¼ ÑÐ²ÑÐ·Ð°Ð½Ð½Ñ‹Ð¹ Ð¾Ð±ÑŠÐµÐºÑ‚ Driver
        .options(joinedload(tb.Order.dispatcher))  # Ð—Ð°Ð³Ñ€ÑƒÐ¶Ð°ÐµÐ¼ ÑÐ²ÑÐ·Ð°Ð½Ð½Ñ‹Ð¹ Ð¾Ð±ÑŠÐµÐºÑ‚ Dispatcher
        .order_by(tb.Order.time)
        .where(tb.Order.idOrder.in_(actiual_order_list))
    )

    
    result = await session.execute(stmt)
    orders = result.unique().scalars().all()    
    formatted_orders = []
    for order in orders:
        cargo_type = await session.scalar(select(tb.CargoType).where(tb.CargoType.idCargoType == order.cargoTypeId))
        formatted_order = await form_order(order=order, cargo_type=cargo_type)
        formatted_orders.append(formatted_order)
    return formatted_orders

@connection
async def get_order(session: AsyncSession, orderId):
    stmt = (
        select(tb.Order)
        .options(joinedload(tb.Order.cargoType))
        .where(tb.Order.idOrder == orderId)
    )
    result = await session.execute(stmt)
    order = result.scalar()

    order.orderTypeName = order.cargoType.cargoTypeName

    session.expunge(order)

    return order

@connection
async def get_cargo_type_name(session: AsyncSession, cargoTypeId):
     cargoType = await session.scalar(select(tb.CargoType).where(tb.CargoType.idCargoType == cargoTypeId))

     return cargoType.cargoTypeName

async def form_order(order, cargo_type, status=None, witoutStatus=False) -> str:
    if hasattr(cargo_type, 'cargoTypeName'):
        cargo_type_name = cargo_type.cargoTypeName  # Ð˜ÑÐ¿Ð¾Ð»ÑŒÐ·ÑƒÐµÐ¼ Ð°Ñ‚Ñ€Ð¸Ð±ÑƒÑ‚
    else:
        cargo_type_name = str(cargo_type)
    # ÐžÑÐ½Ð¾Ð²Ð½Ð¾Ð¹ Ð±Ð»Ð¾Ðº
    formatted_order = [
        hbold(f"ðŸšš Ð—ÐÐšÐÐ— #{order.idOrder}"),
        f"ðŸ“¦ Ð“Ñ€ÑƒÐ·: {order.cargoName}",
        f"ðŸ“ ÐžÐ¿Ð¸ÑÐ°Ð½Ð¸Ðµ: {order.cargoDescription}",
        f"âš–ï¸ Ð’ÐµÑ: {order.cargo_weight} ÐºÐ³",
        f"ðŸ“Œ Ð¢Ð¸Ð¿: {cargo_type_name}",
        f"ðŸ“ ÐžÑ‚Ð¿Ñ€Ð°Ð²Ð»ÐµÐ½Ð¸Ðµ: {order.depart_loc}",
        f"ðŸ Ð”Ð¾ÑÑ‚Ð°Ð²ÐºÐ°: {order.goal_loc}",
        f"ðŸ•’ Ð”Ð°Ñ‚Ð°/Ð²Ñ€ÐµÐ¼Ñ: {order.time.strftime('%d.%m.%Y %H:%M')}",
    ]

    if not witoutStatus:
        status = status or statuses.get(order.orderStatusId)
        formatted_order.append(f"ðŸ”– Ð¡Ñ‚Ð°Ñ‚ÑƒÑ: {hunderline(status)}")

    if order.driverId is not None:
        executors_block = [
            "ðŸ‘¤ ÐžÑ‚Ð²ÐµÑ‚ÑÑ‚Ð²ÐµÐ½Ð½Ñ‹Ðµ:",
            f"ðŸ“ž Ð”Ð¸ÑÐ¿ÐµÑ‚Ñ‡ÐµÑ€: {order.dispatcher.phone}",
            f"ðŸšœ Ð˜ÑÐ¿Ð¾Ð»Ð½Ð¸Ñ‚ÐµÐ»ÑŒ: {order.executor.fio}",
        ]
        formatted_order.extend(executors_block)

    if order.photoId is not None:
        formatted_order.append("ðŸ“¸ Ð¤Ð¾Ñ‚Ð¾ Ð³Ñ€ÑƒÐ·Ð°: Ð¿Ñ€Ð¸Ð»Ð¾Ð¶ÐµÐ½Ð¾")
            

    formatted_order.append('\nâ”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”â”\n')

    return "\n".join(formatted_order)

@connection
async def get_user(session: AsyncSession, tg_id=None, id= None):
    if tg_id != None:
        user = await session.scalar(select(tb.User).where(tb.User.tgId == tg_id))
    if id != None:
        user = await session.scalar(select(tb.User).where(tb.User.idUser == id))
    return user

@connection
async def get_user_role(session: AsyncSession, tg_id):
    user = await session.scalar(select(tb.User).where(tb.User.tgId == tg_id))
    role = await session.scalar(select(tb.Role).where(tb.Role.idRole == user.roleId))
    return role.roleName

@connection
async def chek_next_record(session: AsyncSession, end)-> bool:
    limit = 1
    offset = end - 1
    stmt = select(tb.Order).order_by(tb.Order.time).limit(limit).offset(offset)
    result = await session.execute(stmt)
    order = result.scalar()
    return order is not None

@connection
async def take_order(session: AsyncSession, tg_id, order_id)-> bool:

    if await check_order_status(order_id=order_id, expectStatus = [1]):
        user = await session.scalar(select(tb.User).where(tb.User.tgId == tg_id))
        new_data={
            "driverId": user.idUser,
            "orderStatusId": 2,
            "pickup_time":datetime.now(),
        }

        stmt = (
            update(tb.Order)
            .where(tb.Order.idOrder == order_id)
            .values(**new_data)
        )

        await session.execute(stmt)
        return True
    else:
        return False

@connection
async def check_order_status(session: AsyncSession, order_id, expectStatus: List[int])-> bool:

    order = await session.scalar(select(tb.Order).where(tb.Order.idOrder == order_id))

    return order.orderStatusId in expectStatus

@connection
async def get_order_photo(session: AsyncSession, order_id):

    order = await session.scalar(select(tb.Order).where(tb.Order.idOrder == int(order_id)))
    
    return order.photoId

@connection
async def get_user_for_send(session: AsyncSession, orderId, driver_id, action_text: str):
    order = await session.scalar(select(tb.Order).where(tb.Order.idOrder == orderId))
    disp = await session.scalar(select(tb.User).where(tb.User.idUser == order.dispatcherId))
    driver = await session.scalar(select(tb.User).where(tb.User.tgId == driver_id))
    cargo_type = await session.scalar(select(tb.CargoType).where(tb.CargoType.idCargoType == order.cargoTypeId))
    formatted_order = await form_order(order=order, cargo_type=cargo_type, witoutStatus=True)
    fromatted_mes = (
        f'{action_text}\n'
        f"Ð¢Ñ€Ð°ÑÐ¿Ð¾Ñ€Ñ‚Ð¸Ñ€Ð¾Ð²Ñ‰Ð¸Ðº {driver.fio}\n"
        f"Ð¢ÐµÐ»ÐµÑ„Ð¾Ð½: {driver.phone}\n\n"       
    )
    final_message = fromatted_mes + formatted_order
    return disp.tgId, final_message

@connection
async def get_drivers_for_alarm(session: AsyncSession, order):
    drivers = await session.scalars(select(tb.User).where(tb.User.roleId == 2))

@connection
async def complete_order(session: AsyncSession, tg_id, order_id)-> bool:

    if await check_order_status(order_id=order_id, expectStatus= [2]):
        user = await session.scalar(select(tb.User).where(tb.User.tgId == tg_id))
        new_data={
            "orderStatusId": 3,
            "completion_time":datetime.now(),
        }

        stmt = (
            update(tb.Order)
            .where(tb.Order.idOrder == order_id)
            .values(**new_data)
        )

        await session.execute(stmt)
        return True
    else:
        return False

@connection
async def edit_order(session: AsyncSession, data):
    
    updates = {
        "cargoName": data.get("edit_cargo_name"),
        "cargoDescription": data.get("edit_cargo_description"),
        "cargo_weight": data.get("edit_cargo_weight"),
        "cargoTypeId": data.get("edit_cargo_type_id"),
        "depart_loc": data.get("edit_depart_loc"),
        "goal_loc": data.get("edit_goal_loc"),
        "orderStatusId": data.get("edit_order_status"),
        "isPostponed": data.get("set_postponned")
    }
    if data.get("edit_time") is not None:
        updates["time"] = datetime.strptime(data.get("edit_time"), '%H:%M %d.%m.%Y')
        
    updates = {k: v for k, v in updates.items() if v is not None}
    stmt = (
        update(tb.Order)
        .where(tb.Order.idOrder == int(data["order_id"]))
        .values(**updates)
    )
    await session.execute(stmt)
    
@connection
async def take_off_complete_order(session: AsyncSession, tg_id, order_id)-> None:

        user = await session.scalar(select(tb.User).where(tb.User.tgId == tg_id))
        new_data={
            "orderStatusId": 1,
            "driverId": None,
            "pickup_time": None,
        }

        stmt = (
            update(tb.Order)
            .where(tb.Order.idOrder == order_id)
            .values(**new_data)
        )

        await session.execute(stmt)

FIELDS = {
    "ID Ð·Ð°ÐºÐ°Ð·Ð°": "idOrder",
    "ÐÐ°Ð·Ð²Ð°Ð½Ð¸Ðµ Ð³Ñ€ÑƒÐ·Ð°": "cargoName",
    #"ÐžÐ¿Ð¸ÑÐ°Ð½Ð¸Ðµ Ð³Ñ€ÑƒÐ·Ð°": "cargoDescription",
    "Ð¢Ð¸Ð¿ Ð³Ñ€ÑƒÐ·Ð°": lambda order: order.cargoType.cargoTypeName if order.cargoType else "ÐÐµ ÑƒÐºÐ°Ð·Ð°Ð½",
    #"Ð’ÐµÑ Ð³Ñ€ÑƒÐ·Ð° (ÐºÐ³)": "cargo_weight",
    #"ÐœÐµÑÑ‚Ð¾ Ð¾Ñ‚Ð¿Ñ€Ð°Ð²Ð»ÐµÐ½Ð¸Ñ": "depart_loc",
    #"ÐœÐµÑÑ‚Ð¾ Ð½Ð°Ð·Ð½Ð°Ñ‡ÐµÐ½Ð¸Ñ": "goal_loc",
    "Ð’Ñ€ÐµÐ¼Ñ Ð·Ð°ÐºÐ°Ð·Ð°": lambda order: order.time.strftime("%Y-%m-%d %H:%M:%S"),
    #"Ð¡Ñ‚Ð°Ñ‚ÑƒÑ": lambda order: order.orderStatus.orderStatusName if order.orderStatus else "ÐÐµ ÑƒÐºÐ°Ð·Ð°Ð½",
    "Ð”Ð¸ÑÐ¿ÐµÑ‚Ñ‡ÐµÑ€": lambda order: order.dispatcher.fio if order.dispatcher else "ÐÐµ ÑƒÐºÐ°Ð·Ð°Ð½",
    "Ð’Ð¾Ð´Ð¸Ñ‚ÐµÐ»ÑŒ": lambda order: order.executor.fio if order.executor else "ÐÐµ Ð½Ð°Ð·Ð½Ð°Ñ‡ÐµÐ½",
    "Ð’Ñ€ÐµÐ¼Ñ Ð·Ð°Ð±Ð¾Ñ€Ð°": lambda order: order.pickup_time.strftime("%Y-%m-%d %H:%M:%S") if order.pickup_time else "ÐÐµ ÑƒÐºÐ°Ð·Ð°Ð½Ð¾",
    "Ð’Ñ€ÐµÐ¼Ñ Ð·Ð°Ð²ÐµÑ€ÑˆÐµÐ½Ð¸Ñ": lambda order: order.completion_time.strftime("%Y-%m-%d %H:%M:%S") if order.completion_time else "ÐÐµ ÑƒÐºÐ°Ð·Ð°Ð½Ð¾",
    "Ð’Ñ€ÐµÐ¼Ñ ÑÐ¾Ð·Ð´Ð°Ð½Ð¸Ñ": lambda order: order.create_order_time.strftime("%Y-%m-%d %H:%M:%S"),
    "Ð’Ñ€ÐµÐ¼Ñ Ð²Ñ‹Ð¿Ð¾Ð»Ð½ÐµÐ½Ð¸Ñ Ð·Ð°ÐºÐ°Ð·Ð°": lambda order:( 
        (str(order.completion_time - order.pickup_time)).split('.')[0]
        if order.completion_time and order.pickup_time else None),
    "ÐŸÐµÑ€ÐµÐ½ÐµÑÐµÐ½": lambda order: "Ð”Ð°" if order.isPostponed == True else "ÐÐµÑ‚"
}

@connection 
async def export_orders_to_excel(
    session,
    date_from: datetime = None,
    date_to: datetime = datetime.today() + timedelta(days = 1),
) -> BufferedInputFile:
    
    try:
        # Ð¤Ð¾Ñ€Ð¼Ð¸Ñ€ÑƒÐµÐ¼ Ð·Ð°Ð¿Ñ€Ð¾Ñ Ñ ÑƒÑ‡ÐµÑ‚Ð¾Ð¼ Ñ„Ð¸Ð»ÑŒÑ‚Ñ€Ð¾Ð²
        stmt = (
            select(tb.Order)
            .options(selectinload(tb.Order.cargoType))
            .options(selectinload(tb.Order.executor))
            .options(selectinload(tb.Order.dispatcher))
            .options(selectinload(tb.Order.orderStatus))
        )

        if date_from:
            stmt = stmt.where(tb.Order.create_order_time >= date_from)

        stmt = stmt.where(tb.Order.create_order_time <= date_to).order_by(tb.Order.create_order_time)

        # ÐŸÐ¾Ñ‚Ð¾ÐºÐ¾Ð²Ð°Ñ Ð¾Ð±Ñ€Ð°Ð±Ð¾Ñ‚ÐºÐ° Ð´Ð°Ð½Ð½Ñ‹Ñ… Ð´Ð»Ñ ÑÐºÐ¾Ð½Ð¾Ð¼Ð¸Ð¸ Ð¿Ð°Ð¼ÑÑ‚Ð¸
        result = await session.stream(stmt)

        async def process_data():
            data = []
            async for order in result.scalars():

                order_data = {}
                for display_name, field in FIELDS.items():
                    if callable(field):
                        order_data[display_name] = field(order)
                    else:
                        value = getattr(order, field)
                        order_data[display_name] = value if value is not None else "ÐÐµ ÑƒÐºÐ°Ð·Ð°Ð½Ð¾"
                data.append(order_data)
            if not data:
                raise ValueError("Ð’ Ð±Ð°Ð·Ðµ Ð´Ð°Ð½Ð½Ñ‹Ñ… Ð½ÐµÑ‚ Ð·Ð°ÐºÐ°Ð·Ð¾Ð² Ð´Ð»Ñ ÑƒÐºÐ°Ð·Ð°Ð½Ð½Ñ‹Ñ… Ð¿Ð°Ñ€Ð°Ð¼ÐµÑ‚Ñ€Ð¾Ð²")

            df = pd.DataFrame(data)

            date_column = "Ð’Ñ€ÐµÐ¼Ñ ÑÐ¾Ð·Ð´Ð°Ð½Ð¸Ñ"
            time_to_take_column = "Ð’Ñ€ÐµÐ¼Ñ Ð·Ð°Ð±Ð¾Ñ€Ð°"
            time_to_complete_column = "Ð’Ñ€ÐµÐ¼Ñ Ð·Ð°Ð²ÐµÑ€ÑˆÐµÐ½Ð¸Ñ"
            postponed_column = "ÐŸÐµÑ€ÐµÐ½ÐµÑÐµÐ½"
            
            df[time_to_take_column] = pd.to_datetime(df[time_to_take_column], errors='coerce')
            df[time_to_complete_column] = pd.to_datetime(df[time_to_complete_column], errors='coerce')
            df[date_column] = pd.to_datetime(df[date_column], errors='coerce')
            
            df.sort_values(by=date_column, inplace=True)

            # ÐŸÐ¾Ð»ÑƒÑ‡Ð°ÐµÐ¼ ÑƒÐ½Ð¸ÐºÐ°Ð»ÑŒÐ½Ñ‹Ðµ Ð´Ð°Ñ‚Ñ‹
            dates = df[date_column].dt.date.unique()

            # Ð¡Ð¿Ð¸ÑÐ¾Ðº Ð´Ð»Ñ Ñ…Ñ€Ð°Ð½ÐµÐ½Ð¸Ñ DataFrame'Ð¾Ð² ÐºÐ°Ð¶Ð´Ð¾Ð³Ð¾ Ð´Ð½Ñ
            dfs = []
            summary_rows = []  # ÐÐ¾Ð¼ÐµÑ€Ð° ÑÑ‚Ñ€Ð¾Ðº Ð´Ð»Ñ Ð¾Ð±ÑŠÐµÐ´Ð¸Ð½ÐµÐ½Ð¸Ñ
            current_row = 0
            # ÐžÐ±Ñ€Ð°Ð±Ð°Ñ‚Ñ‹Ð²Ð°ÐµÐ¼ ÐºÐ°Ð¶Ð´Ñ‹Ð¹ Ð´ÐµÐ½ÑŒ
            for date in dates:

                day_df = df[df[date_column].dt.date == date].copy()

                # Ð’Ñ‹Ñ‡Ð¸ÑÐ»ÑÐµÐ¼ Ð¼ÐµÑ‚Ñ€Ð¸ÐºÐ¸ Ð´Ð»Ñ Ð´Ð½Ñ
                avg_time_to_take = (day_df[time_to_take_column] - day_df[date_column]).mean()
                avg_time_to_complete = (day_df[time_to_complete_column] - day_df[time_to_take_column]).mean()
                num_transferred = (day_df[postponed_column] == "Ð”Ð°").sum()
                
                avg_time_to_take_str = str(avg_time_to_take).split(' ')[-1] if pd.notna(avg_time_to_take) else "ÐÐµ ÑƒÐºÐ°Ð·Ð°Ð½Ð¾"
                avg_time_to_complete_str = str(avg_time_to_complete).split(' ')[-1] if pd.notna(avg_time_to_complete) else "ÐÐµ ÑƒÐºÐ°Ð·Ð°Ð½Ð¾"
                # Ð¡Ð¾Ð·Ð´Ð°ÐµÐ¼ ÑÑ‚Ñ€Ð¾ÐºÑƒ ÑÐ²Ð¾Ð´ÐºÐ¸ Ð´Ð»Ñ Ñ‚ÐµÐºÑƒÑ‰ÐµÐ³Ð¾ Ð´Ð½Ñ
                summary_text = (
                            f"Ð”Ð°Ñ‚Ð°: {date.strftime('%d.%m.%Y')} "
                            f"ÑÑ€ÐµÐ´Ð½ÐµÐµ Ð²Ñ€ÐµÐ¼Ñ Ð²Ñ‹Ð¿Ð¾Ð»Ð½ÐµÐ½Ð¸Ñ: {avg_time_to_complete_str.split('.')[0]} "
                            f"ÑÑ€ÐµÐ´Ð½ÐµÐµ Ð²Ñ€ÐµÐ¼Ñ Ð²Ð·ÑÑ‚Ð¸Ñ: {avg_time_to_take_str.split('.')[0]} "
                            f"ÐºÐ¾Ð»Ð¸Ñ‡ÐµÑÑ‚Ð²Ð¾ Ð¿ÐµÑ€ÐµÐ½ÐµÑÐµÐ½Ð½Ñ‹Ñ…: {num_transferred}"
                        )
                summary_df = pd.DataFrame([[summary_text] + [""] * (len(FIELDS) - 1)], columns=df.columns)

                # ÐžÐ±ÑŠÐµÐ´Ð¸Ð½ÑÐµÐ¼ ÑÐ²Ð¾Ð´ÐºÑƒ Ð¸ Ð´Ð°Ð½Ð½Ñ‹Ðµ Ð·Ð°ÐºÐ°Ð·Ð¾Ð² Ð·Ð° Ð´ÐµÐ½ÑŒ
                day_combined = pd.concat([summary_df, day_df], ignore_index=True)

                summary_rows.append(current_row + 1)  # +1 Ð´Ð»Ñ ÑƒÑ‡ÐµÑ‚Ð° Ð·Ð°Ð³Ð¾Ð»Ð¾Ð²ÐºÐ° Ð² Excel
                current_row += len(day_combined)

                dfs.append(day_combined)

            # Ð’Ñ‹Ñ‡Ð¸ÑÐ»ÑÐµÐ¼ Ð¾Ð±Ñ‰Ð¸Ðµ Ð¼ÐµÑ‚Ñ€Ð¸ÐºÐ¸ Ð·Ð° Ð²ÐµÑÑŒ Ð¿ÐµÑ€Ð¸Ð¾Ð´
            overall_avg_time_to_take = (df[time_to_take_column] - df[date_column]).mean()
            overall_avg_time_to_complete = (df[time_to_complete_column] - df[time_to_take_column]).mean()
            overall_num_transferred = (df[postponed_column] == "Ð”Ð°").sum()

            overall_avg_time_to_take_str = str(overall_avg_time_to_take).split(' ')[-1] if pd.notna(overall_avg_time_to_take) else "ÐÐµ ÑƒÐºÐ°Ð·Ð°Ð½Ð¾"
            overall_avg_time_to_complete_str = str(overall_avg_time_to_complete).split(' ')[-1] if pd.notna(overall_avg_time_to_complete) else "ÐÐµ ÑƒÐºÐ°Ð·Ð°Ð½Ð¾"

            # Ð¤Ð¾Ñ€Ð¼Ð°Ñ‚Ð¸Ñ€ÑƒÐµÐ¼ Ð¸Ñ‚Ð¾Ð³Ð¾Ð²ÑƒÑŽ ÑÑ‚Ñ€Ð¾ÐºÑƒ
            overall_summary_text = (
                f"Ð˜Ñ‚Ð¾Ð³Ð¾ "
                f"ÑÑ€ÐµÐ´Ð½ÐµÐµ Ð²Ñ€ÐµÐ¼Ñ Ð²Ñ‹Ð¿Ð¾Ð»Ð½ÐµÐ½Ð¸Ñ: {overall_avg_time_to_complete_str.split('.')[0]} "
                f"ÑÑ€ÐµÐ´Ð½ÐµÐµ Ð²Ñ€ÐµÐ¼Ñ Ð²Ð·ÑÑ‚Ð¸Ñ: {overall_avg_time_to_take_str.split('.')[0]} "
                f"ÐºÐ¾Ð»Ð¸Ñ‡ÐµÑÑ‚Ð²Ð¾ Ð¿ÐµÑ€ÐµÐ½ÐµÑÐµÐ½Ð½Ñ‹Ñ…: {overall_num_transferred}"
            )
            overall_summary_df = pd.DataFrame([[overall_summary_text] + [""] * (len(FIELDS) - 1)], columns=df.columns)
            summary_rows.append(current_row + 1)
            final_df = pd.concat(dfs + [overall_summary_df], ignore_index=True)

            excel_file = BytesIO()
            with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
                final_df.to_excel(writer, index=False, sheet_name='Orders')
                worksheet = writer.sheets['Orders']

                num_columns = len(final_df.columns)
                for row_idx in summary_rows:
                    worksheet.merge_cells(start_row=row_idx + 1, start_column=1, end_row=row_idx + 1, end_column=num_columns)
                    worksheet.cell(row=row_idx + 1, column=1).alignment = Alignment(horizontal='left')

                # Ð¤Ð¾Ñ€Ð¼Ð°Ñ‚Ð¸Ñ€ÑƒÐµÐ¼ ÐºÐ¾Ð»Ð¾Ð½ÐºÐ¸ Ñ Ð²Ñ€ÐµÐ¼ÐµÐ½ÐµÐ¼


                col_idx = list(FIELDS.keys()).index("Ð’Ñ€ÐµÐ¼Ñ Ð²Ñ‹Ð¿Ð¾Ð»Ð½ÐµÐ½Ð¸Ñ Ð·Ð°ÐºÐ°Ð·Ð°") + 1
                for row in range(2, len(final_df) + 2):
                    worksheet.cell(row=row, column=col_idx).number_format = '[h]:mm:ss'

                # ÐÐ°ÑÑ‚Ñ€Ð°Ð¸Ð²Ð°ÐµÐ¼ ÑˆÐ¸Ñ€Ð¸Ð½Ñƒ ÐºÐ¾Ð»Ð¾Ð½Ð¾Ðº
                for idx, column in enumerate(final_df.columns, 1):
                    max_length = max(final_df[column].astype(str).map(len).max(), len(column)) + 2
                    worksheet.column_dimensions[get_column_letter(idx)].width = max_length

            excel_file.seek(0)
            return excel_file.getvalue()

        excel_data = await process_data()

        filename = f"orders_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        file_for_telegram = BufferedInputFile(
            file=excel_data,
            filename=filename
        )

        return file_for_telegram

    except Exception as e:
        logging.error(f"ÐžÑˆÐ¸Ð±ÐºÐ° Ð¿Ñ€Ð¸ ÑÐºÑÐ¿Ð¾Ñ€Ñ‚Ðµ Ð·Ð°ÐºÐ°Ð·Ð¾Ð² Ð² Excel: {e}")
        raise

@connection
async def notificationDrivers(session: AsyncSession,  bot: Bot):

    target_time = datetime.now() + timedelta(minutes=1)
    stmt = (
        select(tb.Order)
        .options(joinedload(tb.Order.executor))
        .options(joinedload(tb.Order.dispatcher))
        .options(joinedload(tb.Order.cargoType))
        .where(and_(
            tb.Order.orderStatusId == 2,
            tb.Order.time >= target_time - timedelta(seconds=30),
            tb.Order.time <= target_time + timedelta(seconds=30)
        ))
    )

    result = await session.execute(stmt)
    orders = result.scalars().all()
    for order in orders:
        mes = "ÐÐ°Ð¿Ð¾Ð¼Ð¸Ð½Ð°Ð½Ð¸Ðµ:\n\n" + await form_order(order=order, cargo_type=order.cargoType.cargoTypeName)
        try:

            await bot.send_message(order.executor.tgId, mes)
        except Exception as e:
            logging.error(f"ÐžÑˆÐ¸Ð±ÐºÐ° Ð¾Ñ‚Ð¿Ñ€Ð°Ð²ÐºÐ¸ ÑÐ¾Ð¾Ð±Ñ‰ÐµÐ½Ð¸Ñ Ð´Ð»Ñ Ð·Ð°ÐºÐ°Ð·Ð° {order.idOrder}: {e}")

@connection
async def dayEnd(session: AsyncSession, bot: Bot):
    stmt = (
        select(tb.Order)
        .options(joinedload(tb.Order.executor))
        .options(joinedload(tb.Order.dispatcher))
        .options(joinedload(tb.Order.cargoType))
        .where(and_(
            tb.Order.orderStatusId == 1,
            tb.Order.time >= datetime.today().replace(hour=0, minute=0, second=0, microsecond=0)
        ))
    )

    orders = (await session.execute(stmt)).scalars().all()
    for order in orders:

        mes = "ÐŸÐµÑ€ÐµÐ½Ð¾Ñ Ð·Ð°ÐºÐ°Ð·Ð°:\n\n" + await form_order(order=order, cargo_type=order.cargoType.cargoTypeName)
        try:

            await bot.send_message(order.dispatcher.tgId, mes, reply_markup= await kb.dayEndKb(orderId=order.idOrder), parse_mode='HTML')
        except Exception as e:
            logging.error(f"ÐžÑˆÐ¸Ð±ÐºÐ° Ð¾Ñ‚Ð¿Ñ€Ð°Ð²ÐºÐ¸ ÑÐ¾Ð¾Ð±Ñ‰ÐµÐ½Ð¸Ñ Ð´Ð»Ñ Ð·Ð°ÐºÐ°Ð·Ð° {order.idOrder}: {e}")

@connection
async def export_diagrama(session, 
    date_from: datetime = None,
    date_to: datetime = datetime.today() + timedelta(days = 1)) -> BufferedInputFile:
    
    stmt = (
        select(tb.Order)
        .options(joinedload(tb.Order.executor))
        .where(and_(
            tb.Order.orderStatusId == 3,
            tb.Order.completion_time >= date_from,
            tb.Order.completion_time <= date_to
        ))
    )

    result = await session.execute(stmt)
    orders = result.scalars().all()

    if not orders:
        raise ValueError("ÐÐµÑ‚ Ð²Ñ‹Ð¿Ð¾Ð»Ð½ÐµÐ½Ð½Ñ‹Ñ… Ð·Ð°ÐºÐ°Ð·Ð¾Ð² Ð·Ð° ÑƒÐºÐ°Ð·Ð°Ð½Ð½Ñ‹Ð¹ Ð¿ÐµÑ€Ð¸Ð¾Ð´")
    
    driver_data = [
        {"Ð’Ð¾Ð´Ð¸Ñ‚ÐµÐ»ÑŒ": order.executor.fio}
        for order in orders if order.executor
    ]

    df = pd.DataFrame(driver_data)
    driver_counts = df["Ð’Ð¾Ð´Ð¸Ñ‚ÐµÐ»ÑŒ"].value_counts()


    plt.figure(figsize=(10, 6))
    bars = plt.bar(driver_counts.index, driver_counts.values, color='skyblue', edgecolor='black')
    plt.title("ÐŸÑ€Ð¾Ð´ÑƒÐºÑ‚Ð¸Ð²Ð½Ð¾ÑÑ‚ÑŒ Ð²Ð¾Ð´Ð¸Ñ‚ÐµÐ»ÐµÐ¹")
    plt.xlabel("Ð’Ð¾Ð´Ð¸Ñ‚ÐµÐ»ÑŒ")
    plt.ylabel("ÐšÐ¾Ð»Ð¸Ñ‡ÐµÑÑ‚Ð²Ð¾ Ð·Ð°ÐºÐ°Ð·Ð¾Ð²")
    plt.xticks(rotation=45, ha='right')
    plt.grid(True)

    for bar in bars:
        yval = bar.get_height()
        plt.text(bar.get_x() + bar.get_width()/2, yval + 0.1, int(yval), ha='center', va='bottom')

    hist_file = BytesIO()
    plt.savefig(hist_file, format='png', bbox_inches='tight')
    plt.close()
    hist_file.seek(0)

    hist_filename = f"Ð”Ð¸Ð°Ð³Ñ€Ð°Ð¼Ð¼Ð°_Ñ‚Ñ€Ð°Ð½ÑÐ¿Ð¾Ñ€Ñ‚Ð¸Ñ€Ð¾Ð²Ñ‰Ð¸ÐºÐ¾Ð²_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png"
    return BufferedInputFile(file=hist_file.getvalue(), filename=hist_filename)

@connection
async def get_user_id(session: AsyncSession, tg_id: int) -> int:
    if tg_id in user_cache:  
        return user_cache[tg_id]  
    
    user = await session.scalar(select(tb.User).where(tb.User.tgId == tg_id))
    user_cache[tg_id] = user.idUser
    return user.idUser

@connection
async def save_location(session: AsyncSession, user_id, latitude, longitude, timestamp):

    new_loc = tb.UserLocation(
        user_id=user_id,
        latitude=latitude,
        longitude=longitude,
        timestamp=timestamp,
    )

    session.add(new_loc)

@connection
async def get_map(session: AsyncSession, tg_id, date):
    
    #user_id = await get_user_id(tg_id = tg_id)
    user_id = 2
    date = datetime.today() - timedelta(days=1)

    stmt = (
        select(tb.UserLocation)
        .where(and_ (func.date(tb.UserLocation.timestamp) == date.date(),
                     tb.UserLocation.user_id == user_id))
        .order_by((tb.UserLocation.created_at))
    )

    results = await session.stream(stmt)
    locations = await results.scalars().all()

    if len(locations) < 2:
        return "ÐÐµÐ´Ð¾ÑÑ‚Ð°Ñ‚Ð¾Ñ‡Ð½Ð¾ Ð´Ð°Ð½Ð½Ñ‹Ñ… Ð´Ð»Ñ Ð¿Ð¾ÑÑ‚Ñ€Ð¾ÐµÐ½Ð¸Ñ Ð¼Ð°Ñ€ÑˆÑ€ÑƒÑ‚Ð°. Ð¢Ñ€ÐµÐ±ÑƒÐµÑ‚ÑÑ Ñ…Ð¾Ñ‚Ñ Ð±Ñ‹ Ð´Ð²Ðµ Ñ‚Ð¾Ñ‡ÐºÐ¸."
    
    coordinates = [(loc.longitude, loc.latitude) for loc in locations]
    coordinates = coordinates[::10]
    route_points = ','.join([f'{lat},{lon}' for lat, lon in coordinates])

    loop = asyncio.get_event_loop()
    image_data = await loop.run_in_executor(None, generate_map, coordinates)

    return BufferedInputFile(file=image_data, filename="route_map.png")

def generate_map(coordinates):
    
    
    # Ð¡Ð¾Ð·Ð´Ð°Ñ‘Ð¼ Ð¾Ð±ÑŠÐµÐºÑ‚ ÐºÐ°Ñ€Ñ‚Ñ‹ Ñ Ñ€Ð°Ð·Ð¼ÐµÑ€Ð°Ð¼Ð¸ 800x600 Ð¿Ð¸ÐºÑÐµÐ»ÐµÐ¹
    m = StaticMap(800, 600)
    
    # Ð¡Ð¾Ð·Ð´Ð°Ñ‘Ð¼ Ð»Ð¸Ð½Ð¸ÑŽ Ð¼Ð°Ñ€ÑˆÑ€ÑƒÑ‚Ð°: ÑÐ¸Ð½ÑÑ, Ñ‚Ð¾Ð»Ñ‰Ð¸Ð½Ð¾Ð¹ 5 Ð¿Ð¸ÐºÑÐµÐ»ÐµÐ¹
    line = Line(coordinates, 'blue', 5)
    m.add_line(line)

    # Ð ÐµÐ½Ð´ÐµÑ€Ð¸Ð¼ ÐºÐ°Ñ€Ñ‚Ñƒ Ð² Ð¸Ð·Ð¾Ð±Ñ€Ð°Ð¶ÐµÐ½Ð¸Ðµ
    image = m.render()

    # Ð¡Ð¾Ñ…Ñ€Ð°Ð½ÑÐµÐ¼ Ð¸Ð·Ð¾Ð±Ñ€Ð°Ð¶ÐµÐ½Ð¸Ðµ Ð² Ð±ÑƒÑ„ÐµÑ€ Ð¿Ð°Ð¼ÑÑ‚Ð¸
    image_io = BytesIO()
    image.save(image_io, 'PNG')
    image_io.seek(0)

    return image_io.getvalue()

@connection
async def change_role(session: AsyncSession, data, id_role):
        
        user = await session.scalar(select(tb.User).where(tb.User.tgId == data["tg_id"]))
        new_data={
            "roleId": id_role
        }

        stmt = (
            update(tb.User)
            .where(tb.User.idUser == user.idUser)
            .values(**new_data)
        )

        await session.execute(stmt)