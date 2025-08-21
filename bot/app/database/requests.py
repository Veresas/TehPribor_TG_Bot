import numpy as np
from app.database.models import async_session
import app.database.models as tb
from sqlalchemy import select, and_, update, func
from sqlalchemy.orm import joinedload, selectinload
import logging
from datetime import datetime, timedelta
from aiogram.utils.markdown import hbold, hunderline
from aiogram.types import BufferedInputFile
from aiogram import Bot
from typing import List
import pandas as pd
from io import BytesIO
from openpyxl.utils import get_column_letter
import app.keyboards as kb
from sqlalchemy.ext.asyncio import AsyncSession
from openpyxl.styles import Alignment
import matplotlib.pyplot as plt
from cachetools import TTLCache

from staticmap import StaticMap, Line
import asyncio

user_cache = TTLCache(maxsize=100, ttl=300)
dep_build_cache = TTLCache(maxsize=200, ttl=86400)

def connection(func):
    async def inner(*args, **kwargs):
        async with async_session() as session:
            try:
                result = await func(session, *args, **kwargs)
                await session.commit()
                return result
            except Exception as e:
                await session.rollback()
                logging.error(f"Ошибка в функции {func.__name__}: {e}")
                raise e
            finally:
                await session.close()
    return inner

@connection
async def check_user(session: AsyncSession, tg_id)-> bool:
    user = await session.scalar(select(tb.User).where(tb.User.tgId == tg_id))

    return user is not None


@connection
async def reg_user(session: AsyncSession, data, tg_id)-> int:
    tg_id = int(tg_id)
    new_user = tb.User(
        tgId = tg_id,
        phone=data.get("number"),
        fio=data.get("fio"),
        role_id = None,
        is_denied = True
    )

    session.add(new_user)

@connection
async def add_role_and_acess(session: AsyncSession, tgId, role):
    role = await session.scalar(select(tb.Role).where(tb.Role.role_name == role))
    new_data={
        "role_id": role.id_role,
        "is_denied": False
    }

    stmt = (
        update(tb.User)
        .where(tb.User.tgId == int(tgId))
        .values(**new_data)
    )

    await session.execute(stmt)

@connection
async def get_cargo_types(session: AsyncSession):
    result = await session.execute(select(tb.CargoType).order_by(tb.CargoType.cargo_type_name))
    cargo_types = result.scalars().all()
    return {cargo.id_cargo_type: cargo.cargo_type_name for cargo in cargo_types}

@connection
async def get_cargo_type_name_by_id(session: AsyncSession, data):
    cargo_type_name = await session.scalar(select(tb.CargoType).where(tb.CargoType.id_cargo_type == data))

    return cargo_type_name.cargo_type_name



@connection
async def add_new_order(session: AsyncSession, data):
    disp_id = await session.scalar(select(tb.User).where(tb.User.tgId == data["tg_id"]))
    new_order = tb.Order(
        cargo_name=data["cargo_name"],
        cargo_description=data["cargo_description"],
        cargo_type_id=int(data["cargo_type_id"]),
        cargo_weight=float(data["cargo_weight"]),
        depart_loc=data["depart_loc_id"],
        goal_loc=data["goal_loc_id"],
        time=datetime.strptime(data["time"], '%H:%M %d.%m.%Y'),
        order_status_id = 1,
        dispatcher_id = disp_id.id_user,
        driver_id=None,
        create_order_time = datetime.now(),
    )
    if "photoId" in data:
        new_order.photo_id = data["photoId"]
    if "isUrgent" in data:
        new_order.is_urgent = data["isUrgent"]

    session.add(new_order)
    await session.flush()
    await session.refresh(new_order)
    return new_order.id_order

@connection
async def alarm_for_drivers(session: AsyncSession, orderId, bot: Bot):
    drivers = await session.scalars(select(tb.User).where(tb.User.role_id == 2))
    order = await session.scalar(select(tb.Order).options(joinedload(tb.Order.cargoType)).where(tb.Order.id_order == orderId))
    mes = f'Срочный заказ:\n\n' + await form_order(order=order, cargo_type=order.cargoType.cargo_type_name)
    for driver in drivers:
        if (driver.tgId != None):
            await bot.send_message(driver.tgId, mes, reply_markup=await kb.alarm_kb(orderId=orderId), parse_mode="HTML")

statuses = {
    1: "Доступен",
    2: "В работе",
    3: "Завершен",
    4: "Отменен"
}

@connection
async def get_order_keys(session: AsyncSession, dateTime: datetime = None, tg_id = None, isActual = False, isPrivateCatalog =False, statusId:int = None):
    stmt = select(tb.Order)
    user = await session.scalar(select(tb.User).where(tb.User.tgId == tg_id))
    if not isPrivateCatalog:
        if(dateTime != None):
            start_time = dateTime
            end_time = dateTime + timedelta(days=1)
            stmt = stmt.where(and_(tb.Order.time > start_time,
                                    tb.Order.time < end_time))
        if user.role_id == 2:
            stmt = stmt.where(tb.Order.order_status_id == 1)
        stmt = stmt.order_by(tb.Order.time)

    else:
        match user.role_id:
            case 1: #Диспетчер
                role_condition= tb.Order.dispatcher_id == user.id_user
                
            case 2:  # Водитель
                role_condition = tb.Order.driver_id == user.id_user
            case 3:
                role_condition= tb.Order.dispatcher_id == user.id_user
            case 4:
                role_condition= tb.Order.dispatcher_id == user.id_user
            case _:
                raise ValueError(f"Роль {user.role_id} не поддерживается")
        
        stmt = stmt.where(role_condition)

        if isActual:
            if statusId in [1,2]:
                stmt = stmt.where(tb.Order.order_status_id == statusId)
            else:
                stmt = stmt.where(tb.Order.order_status_id.in_([1,2]))
            statusId = None
        
        stmt = stmt.order_by(tb.Order.time.desc())
    
    if statusId is not None:

        stmt = stmt.where(tb.Order.order_status_id == statusId)    

    result = await session.execute(stmt)
    orders = result.scalars().all()    

    
    order_keys = []
    for order in orders:
        order_keys.append(order.id_order)
    return order_keys

@connection
async def get_orders(session: AsyncSession, ordersKeys, start: int, end: int):
    actiual_order_list = ordersKeys[start:end]
    stmt = (
        select(tb.Order)
        .options(joinedload(tb.Order.executor))  # Загружаем связанный объект Driver
        .options(joinedload(tb.Order.dispatcher))  # Загружаем связанный объект Dispatcher
        .order_by(tb.Order.time)
        .where(tb.Order.id_order.in_(actiual_order_list))
    )

    
    result = await session.execute(stmt)
    orders = result.unique().scalars().all()    
    formatted_orders = []
    for order in orders:
        cargo_type = await session.scalar(select(tb.CargoType).where(tb.CargoType.id_cargo_type == order.cargo_type_id))
        formatted_order = await form_order(order=order, cargo_type=cargo_type)
        formatted_orders.append(formatted_order)
    return formatted_orders

@connection
async def get_order(session: AsyncSession, orderId):
    stmt = (
        select(tb.Order)
        .options(joinedload(tb.Order.cargoType))
        .where(tb.Order.id_order == orderId)
    )
    result = await session.execute(stmt)
    order = result.scalar()

    order.orderTypeName = order.cargoType.cargo_type_name

    session.expunge(order)

    return order

@connection
async def get_cargo_type_name(session: AsyncSession, cargoTypeId):
     cargoType = await session.scalar(select(tb.CargoType).where(tb.CargoType.id_cargo_type == cargoTypeId))

     return cargoType.cargo_type_name

async def form_order(order, cargo_type, status=None, witoutStatus=False) -> str:
    if hasattr(cargo_type, 'cargo_type_name'):
        cargo_type_name = cargo_type.cargo_type_name  # Используем атрибут
    else:
        cargo_type_name = str(cargo_type)
    # Основной блок
    formatted_order = [
        hbold(f"🚚 ЗАКАЗ #{order.id_order}"),
        f"📦 Груз: {order.cargo_name}",
        f"📝 Описание: {order.cargo_description}",
        f"⚖️ Вес: {order.cargo_weight} кг",
        f"📌 Тип: {cargo_type_name}",
        f"📍 Отправление: {get_dep_build_input(order.depart_loc)}",
        f"🏁 Доставка: {get_dep_build_input(order.goal_loc)}",
        f"🕒 Дата/время: {order.time.strftime('%d.%m.%Y %H:%M')}",
    ]

    if not witoutStatus:
        status = status or statuses.get(order.order_status_id)
        formatted_order.append(f"🔖 Статус: {hunderline(status)}")

    if order.driver_id is not None:
        executors_block = [
            "👤 Ответственные:",
            f"📞 Диспетчер: {order.dispatcher.phone}",
            f"🚜 Исполнитель: {order.executor.fio}",
        ]
        formatted_order.extend(executors_block)

    if order.photo_id is not None:
        formatted_order.append("📸 Фото груза: приложено")
            

    formatted_order.append('\n━━━━━━━━━━━━━━━━━━\n')

    return "\n".join(formatted_order)

@connection
async def get_user(session: AsyncSession, tg_id=None, id= None):
    if tg_id != None:
        user = await session.scalar(select(tb.User).where(tb.User.tgId == tg_id))
    if id != None:
        user = await session.scalar(select(tb.User).where(tb.User.id_user == id))
    if user is None:
        return None
    
    await session.refresh(
        user,
        attribute_names=[
            "id_user",
            "tgId",
            "phone",
            "fio",
            "role_id",
            "is_denied"
        ]
    )
    session.expunge(user)
    return user

@connection
async def get_user_role(session: AsyncSession, tg_id):
    tg_id = int(tg_id)
    user = await session.scalar(select(tb.User).where(tb.User.tgId == tg_id))
    role = await session.scalar(select(tb.Role).where(tb.Role.id_role == user.role_id))
    if role == None:
        return None
    return role.role_name

@connection
async def chek_next_record(session: AsyncSession, end)-> bool:
    limit = 1
    offset = end - 1
    stmt = select(tb.Order).order_by(tb.Order.time).limit(limit).offset(offset)
    result = await session.execute(stmt)
    order = result.scalar()
    return order is not None

@connection
async def take_order(session: AsyncSession, tg_id, order_id) -> str:
    user = await session.scalar(select(tb.User).where(tb.User.tgId == tg_id))
    if await check_order_status(order_id=order_id, expectStatus=[1]):
        new_data = {
            "driver_id": user.id_user,
            "order_status_id": 2,
            "pickup_time": datetime.now(),
        }
        stmt = (
            update(tb.Order)
            .where(tb.Order.id_order == order_id)
            .values(**new_data)
        )
        await session.execute(stmt)
        return 'ok'
    else:
        return 'taken'

@connection
async def check_order_status(session: AsyncSession, order_id, expectStatus: List[int])-> bool:

    order = await session.scalar(select(tb.Order).where(tb.Order.id_order == order_id))

    return order.order_status_id in expectStatus

@connection
async def get_order_photo(session: AsyncSession, order_id):

    order = await session.scalar(select(tb.Order).where(tb.Order.id_order == int(order_id)))
    
    return order.photo_id

@connection
async def get_user_for_send(session: AsyncSession, orderId, driver_id, action_text: str, optin_mes: str = None):
    order = await session.scalar(select(tb.Order).where(tb.Order.id_order == orderId))
    disp = await session.scalar(select(tb.User).where(tb.User.id_user == order.dispatcher_id))
    driver = await session.scalar(select(tb.User).where(tb.User.tgId == driver_id))
    cargo_type = await session.scalar(select(tb.CargoType).where(tb.CargoType.id_cargo_type == order.cargo_type_id))
    formatted_order = await form_order(order=order, cargo_type=cargo_type, witoutStatus=True)
    fromatted_mes = (
        f'{action_text}\n'
        f"Траспортировщик {driver.fio}\n"
        f"Телефон: {driver.phone}\n\n"       
    )
    final_message = fromatted_mes + formatted_order
    if optin_mes != None:
        final_message = final_message + optin_mes
    return disp.tgId, final_message

@connection
async def get_drivers_for_alarm(session: AsyncSession, order):
    drivers = await session.scalars(select(tb.User).where(tb.User.role_id == 2).where(tb.User.tgId != None))

@connection
async def complete_order(session: AsyncSession, tg_id, order_id)-> bool:

    if await check_order_status(order_id=order_id, expectStatus= [2]):
        user = await session.scalar(select(tb.User).where(tb.User.tgId == tg_id))
        new_data={
            "order_status_id": 3,
            "completion_time":datetime.now(),
        }

        stmt = (
            update(tb.Order)
            .where(tb.Order.id_order == order_id)
            .values(**new_data)
        )

        await session.execute(stmt)
        return True
    else:
        return False

@connection
async def edit_order(session: AsyncSession, data):
    
    updates = {
        "cargo_name": data.get("edit_cargo_name"),
        "cargo_description": data.get("edit_cargo_description"),
        "cargo_weight": data.get("edit_cargo_weight"),
        "cargo_type_id": data.get("edit_cargo_type_id"),
        "depart_loc": data.get("edit_depart_loc"),
        "goal_loc": data.get("edit_goal_loc"),
        "order_status_id": data.get("edit_order_status"),
        "is_postponed": data.get("set_postponned")
    }
    if data.get("edit_time") is not None:
        updates["time"] = datetime.strptime(data.get("edit_time"), '%H:%M %d.%m.%Y')
        
    updates = {k: v for k, v in updates.items() if v is not None}
    stmt = (
        update(tb.Order)
        .where(tb.Order.id_order == int(data["order_id"]))
        .values(**updates)
    )
    await session.execute(stmt)
    
@connection
async def take_off_complete_order(session: AsyncSession, tg_id, order_id)-> None:

        user = await session.scalar(select(tb.User).where(tb.User.tgId == tg_id))
        new_data={
            "order_status_id": 1,
            "driver_id": None,
            "pickup_time": None,
        }

        stmt = (
            update(tb.Order)
            .where(tb.Order.id_order == order_id)
            .values(**new_data)
        )

        await session.execute(stmt)

FIELDS = {
    "ID заказа": "id_order",
    "Название груза": "cargo_name",
    #"Описание груза": "cargo_description",
    "Тип груза": lambda order: order.cargoType.cargo_type_name if order.cargoType else "Не указан",
    #"Вес груза (кг)": "cargo_weight",
    "Место отправления": lambda order: f'Отдел: {order.depart_loc_ref.department.department_name}, корпус: {order.depart_loc_ref.building.building_name}'
    if order.depart_loc_ref and order.depart_loc_ref.department and order.depart_loc_ref.building else "Не указано",
    "Место назначения": lambda order: f'Отдел: {order.goal_loc_ref.department.department_name}, корпус: {order.goal_loc_ref.building.building_name}'
    if order.goal_loc_ref and order.goal_loc_ref.department and order.goal_loc_ref.building else "Не указано",
    "Время заказа": lambda order: order.time.strftime("%Y-%m-%d %H:%M:%S"),
    #"Статус": lambda order: order.orderStatus.order_status_name if order.orderStatus else "Не указан",
    "Диспетчер": lambda order: order.dispatcher.fio if order.dispatcher else "Не указан",
    "Водитель": lambda order: order.executor.fio if order.executor else "Не назначен",
    "Время забора": lambda order: order.pickup_time.strftime("%Y-%m-%d %H:%M:%S") if order.pickup_time else "Не указано",
    "Время завершения": lambda order: order.completion_time.strftime("%Y-%m-%d %H:%M:%S") if order.completion_time else "Не указано",
    "Время создания": lambda order: order.create_order_time.strftime("%Y-%m-%d %H:%M:%S"),
    "Время выполнения заказа": lambda order:( 
        (str(order.completion_time - order.pickup_time)).split('.')[0]
        if order.completion_time and order.pickup_time else None),
    "Перенесен": lambda order: "Да" if order.is_postponed == True else "Нет"
}

@connection 
async def export_orders_to_excel(
    session,
    date_from: datetime = None,
    date_to: datetime = datetime.today() + timedelta(days = 1),
) -> BufferedInputFile:
    
    try:
        # Формируем запрос с учетом фильтров
        stmt = (
            select(tb.Order)
            .options(selectinload(tb.Order.cargoType))
            .options(selectinload(tb.Order.executor))
            .options(selectinload(tb.Order.dispatcher))
            .options(selectinload(tb.Order.orderStatus))
            .options(
                        joinedload(tb.Order.depart_loc_ref)
                        .joinedload(tb.DepartmentBuilding.department)
                    )
            .options(
                        joinedload(tb.Order.depart_loc_ref)
                        .joinedload(tb.DepartmentBuilding.building)
                    )
            .options(
                        joinedload(tb.Order.goal_loc_ref)
                        .joinedload(tb.DepartmentBuilding.department)
                    )
            .options(
                        joinedload(tb.Order.goal_loc_ref)
                        .joinedload(tb.DepartmentBuilding.building)
                    )
        )

        if date_from:
            stmt = stmt.where(tb.Order.create_order_time >= date_from)

        stmt = stmt.where(tb.Order.create_order_time <= date_to).order_by(tb.Order.create_order_time)

        # Потоковая обработка данных для экономии памяти
        result = await session.stream(stmt)

        async def process_data():
            data = []
            async for order in result.scalars():

                order_data = {}
                for display_name, field in FIELDS.items():
                    if callable(field):
                        order_data[display_name] = field(order)
                    else:
                        value = getattr(order, field)
                        order_data[display_name] = value if value is not None else "Не указано"
                data.append(order_data)
            if not data:
                raise ValueError("В базе данных нет заказов для указанных параметров")

            df = pd.DataFrame(data)

            date_column = "Время создания"
            time_to_take_column = "Время забора"
            time_to_complete_column = "Время завершения"
            postponed_column = "Перенесен"
            
            #Попроавака на часовой пояс москвы +3 добавлен + pd.Timedelta(hours=3)
            df[time_to_take_column] = pd.to_datetime(df[time_to_take_column], errors='coerce') + pd.Timedelta(hours=3)
            df[time_to_complete_column] = pd.to_datetime(df[time_to_complete_column], errors='coerce') + pd.Timedelta(hours=3)
            df[date_column] = pd.to_datetime(df[date_column], errors='coerce') + pd.Timedelta(hours=3)
            
            df.sort_values(by=date_column, inplace=True)

            # Получаем уникальные даты
            dates = df[date_column].dt.date.unique()

            # Список для хранения DataFrame'ов каждого дня
            dfs = []
            summary_rows = []  # Номера строк для объединения
            current_row = 0
            # Обрабатываем каждый день
            for date in dates:

                day_df = df[df[date_column].dt.date == date].copy()

                # Вычисляем метрики для дня
                avg_time_to_take = (day_df[time_to_take_column] - day_df[date_column]).mean()
                avg_time_to_complete = (day_df[time_to_complete_column] - day_df[time_to_take_column]).mean()
                num_transferred = (day_df[postponed_column] == "Да").sum()
                
                avg_time_to_take_str = str(avg_time_to_take).split(' ')[-1] if pd.notna(avg_time_to_take) else "Не указано"
                avg_time_to_complete_str = str(avg_time_to_complete).split(' ')[-1] if pd.notna(avg_time_to_complete) else "Не указано"
                # Создаем строку сводки для текущего дня
                summary_text = (
                            f"Дата: {date.strftime('%d.%m.%Y')} "
                            f"среднее время выполнения: {avg_time_to_complete_str.split('.')[0]} "
                            f"среднее время взятия: {avg_time_to_take_str.split('.')[0]} "
                            f"количество перенесенных: {num_transferred}"
                        )
                summary_df = pd.DataFrame([[summary_text] + [""] * (len(FIELDS) - 1)], columns=df.columns)

                # Объединяем сводку и данные заказов за день
                day_combined = pd.concat([summary_df, day_df], ignore_index=True)

                summary_rows.append(current_row + 1)  # +1 для учета заголовка в Excel
                current_row += len(day_combined)

                dfs.append(day_combined)

            # Вычисляем общие метрики за весь период
            overall_avg_time_to_take = (df[time_to_take_column] - df[date_column]).mean()
            overall_avg_time_to_complete = (df[time_to_complete_column] - df[time_to_take_column]).mean()
            overall_num_transferred = (df[postponed_column] == "Да").sum()

            overall_avg_time_to_take_str = str(overall_avg_time_to_take).split(' ')[-1] if pd.notna(overall_avg_time_to_take) else "Не указано"
            overall_avg_time_to_complete_str = str(overall_avg_time_to_complete).split(' ')[-1] if pd.notna(overall_avg_time_to_complete) else "Не указано"

            # Форматируем итоговую строку
            overall_summary_text = (
                f"Итого "
                f"среднее время выполнения: {overall_avg_time_to_complete_str.split('.')[0]} "
                f"среднее время взятия: {overall_avg_time_to_take_str.split('.')[0]} "
                f"количество перенесенных: {overall_num_transferred}"
            )
            overall_summary_df = pd.DataFrame([[overall_summary_text] + [""] * (len(FIELDS) - 1)], columns=df.columns)
            summary_rows.append(current_row + 1)
            final_df = pd.concat(dfs + [overall_summary_df], ignore_index=True)

            excel_file = BytesIO()
            with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
                final_df.to_excel(writer, index=False, sheet_name='Orders')
                worksheet = writer.sheets['Orders']

                num_columns = len(final_df.columns)
                for row_idx in summary_rows:
                    worksheet.merge_cells(start_row=row_idx + 1, start_column=1, end_row=row_idx + 1, end_column=num_columns)
                    worksheet.cell(row=row_idx + 1, column=1).alignment = Alignment(horizontal='left')

                # Форматируем колонки с временем


                col_idx = list(FIELDS.keys()).index("Время выполнения заказа") + 1
                for row in range(2, len(final_df) + 2):
                    worksheet.cell(row=row, column=col_idx).number_format = '[h]:mm:ss'

                # Настраиваем ширину колонок
                for idx, column in enumerate(final_df.columns, 1):
                    max_length = max(final_df[column].astype(str).map(len).max(), len(column)) + 2
                    worksheet.column_dimensions[get_column_letter(idx)].width = max_length

            excel_file.seek(0)
            return excel_file.getvalue()

        excel_data = await process_data()

        filename = f"orders_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        file_for_telegram = BufferedInputFile(
            file=excel_data,
            filename=filename
        )

        return file_for_telegram

    except Exception as e:
        logging.error(f"Ошибка при экспорте заказов в Excel: {e}")
        raise

@connection
async def notificationDrivers(session: AsyncSession,  bot: Bot):

    target_time = datetime.now() + timedelta(minutes=1)
    stmt = (
        select(tb.Order)
        .options(joinedload(tb.Order.executor))
        .options(joinedload(tb.Order.dispatcher))
        .options(joinedload(tb.Order.cargoType))
        .where(and_(
            tb.Order.order_status_id == 2,
            tb.Order.time >= target_time - timedelta(seconds=30),
            tb.Order.time <= target_time + timedelta(seconds=30)
        ))
    )

    result = await session.execute(stmt)
    orders = result.scalars().all()
    for order in orders:
        mes = "Напоминание:\n\n" + await form_order(order=order, cargo_type=order.cargoType.cargo_type_name)
        try:
            if (order.executor.tgId != None):
                await bot.send_message(order.executor.tgId, mes)
        except Exception as e:
            logging.error(f"Ошибка отправки сообщения для заказа {order.id_order}: {e}")

@connection
async def notiNewOrders(session: AsyncSession,  bot: Bot):
    target_time = datetime.now().replace(hour=0,minute=0,second=0)
    stmt = (
        select(tb.Order)
        .where(and_(
            tb.Order.order_status_id == 1,
            tb.Order.create_order_time > target_time
        ))
    )
    result = await session.execute(stmt)
    orders = result.scalars().all()
    countOrders = len(orders)
    if countOrders > 0:
        drivers = await session.scalars(select(tb.User).where(tb.User.role_id == 2))
        drivers = drivers.all()
        for driver in drivers:
            try:
                if (driver.tgId != None):
                    await bot.send_message(driver.tgId, f'Доступно {countOrders} заказ(ов)')
            except Exception as e:
                logging.error(f"Ошибка отправки сообщения для водителя {driver.tgId}: {e}")


@connection
async def dayEnd(session: AsyncSession, bot: Bot):
    stmt = (
        select(tb.Order)
        .options(joinedload(tb.Order.executor))
        .options(joinedload(tb.Order.dispatcher))
        .options(joinedload(tb.Order.cargoType))
        .where(and_(
            tb.Order.order_status_id == 1,
            tb.Order.time >= datetime.today().replace(hour=0, minute=0, second=0, microsecond=0)
        ))
    )

    orders = (await session.execute(stmt)).scalars().all()
    for order in orders:

        mes = "Перенос заказа:\n\n" + await form_order(order=order, cargo_type=order.cargoType.cargo_type_name)
        try:

            await bot.send_message(order.dispatcher.tgId, mes, reply_markup= await kb.dayEndKb(orderId=order.id_order), parse_mode='HTML')
        except Exception as e:
            logging.error(f"Ошибка отправки сообщения для заказа {order.id_order}: {e}")

def create_figure_with_subplots(n_subplots=1, figsize=(17, 19), height_ratios=None):
    """Создает фигуру с подграфиками"""
    if height_ratios is None:
        height_ratios = [1] * n_subplots
    fig, axes = plt.subplots(n_subplots, 1, figsize=figsize, height_ratios=height_ratios)
    plt.subplots_adjust(left=0.1, right=0.65, top=0.95, bottom=0.1, hspace=0.3)
    return fig, axes

def save_figure_to_buffer(fig):
    """Сохраняет фигуру в буфер"""
    buffer = BytesIO()
    fig.savefig(buffer, format='png', bbox_inches='tight')
    plt.close(fig)
    buffer.seek(0)
    return buffer

def create_driver_stats(orders):
    """Создает статистику по водителям"""
    data_to_time = [
        {
            "Водитель": order.executor.fio,
            "Время выполнения (сек)": (
                (order.completion_time - order.pickup_time).total_seconds()
                if order.completion_time.date() == order.pickup_time.date()
                else None
            )
        }
        for order in orders 
        if order.executor and order.cargoType and order.completion_time.date() == order.pickup_time.date()
        and getattr(order.executor, 'is_denied', True) == False and getattr(order.executor, 'role_id', None) == 2
    ]

    data_to_order_count = [
        {
            "Водитель": order.executor.fio,
            "Группа груза": order.cargoType.cargo_type_name
        }
        for order in orders 
        if order.executor and order.cargoType
        and getattr(order.executor, 'is_denied', True) == False and getattr(order.executor, 'role_id', None) == 2
    ]

    df_time = pd.DataFrame(data_to_time)
    df_orders = pd.DataFrame(data_to_order_count)
    
    return {
        'driver_time': df_time.groupby("Водитель")["Время выполнения (сек)"].mean(),
        'driver_cargo': pd.crosstab(df_orders["Водитель"], df_orders["Группа груза"]),
        'driver_counts': df_orders["Водитель"].value_counts(),
        'cargo_counts': df_orders["Группа груза"].value_counts()
    }

def plot_driver_stats(ax1, ax2, stats):
    """Строит графики статистики водителей"""
    # График распределения заказов по типам грузов
    stats['driver_cargo'].plot(
        kind='bar', 
        stacked=True, 
        ax=ax1, 
        color=plt.cm.Set3(range(len(stats['driver_cargo'].columns))), 
        edgecolor='black'
    )
    ax1.set_title("Количество заказов по водителям с разбиением по типам грузов")
    ax1.set_xlabel("Водитель")
    ax1.set_ylabel("Количество заказов")
    ax1.tick_params(axis='x', rotation=45)
    ax1.grid(True, axis='y')
    ax1.legend(title="Группа груза", bbox_to_anchor=(1.05, 1), loc='upper left')

    # Добавляем значения на столбцы
    for i, driver in enumerate(stats['driver_cargo'].index):
        cumulative_height = 0
        for j, cargo_type in enumerate(stats['driver_cargo'].columns):
            value = stats['driver_cargo'].loc[driver, cargo_type]
            if value > 0:
                cumulative_height += value
                text_y = cumulative_height - (value / 2)
                ax1.text(i, text_y, int(value), ha='center', va='center', fontsize=8, color='black')

        total = stats['driver_cargo'].sum(axis=1)[driver]
        ax1.text(i, total + 0.5, int(total), ha='center', va='bottom')

    # График среднего времени выполнения
    bars = ax2.bar(
        stats['driver_time'].index, 
        stats['driver_time'].values / 60,  # Переводим секунды в минуты
        color='lightgreen', 
        edgecolor='black'
    )
    ax2.set_title("Среднее время выполнения заказов по водителям (в минутах)")
    ax2.set_xlabel("Водитель")
    ax2.set_ylabel("Среднее время (мин)")
    ax2.tick_params(axis='x', rotation=45)
    ax2.grid(True, axis='y')
    
    for bar in bars:
        yval = bar.get_height()
        ax2.text(bar.get_x() + bar.get_width()/2, yval + 0.5, f"{yval:.1f}", ha='center', va='bottom')

def create_legend_table(fig, cargo_counts, driver_counts):
    """Создает таблицу с легендой"""
    ax_table = fig.add_subplot(111)
    ax_table.axis('off')

    # Таблица по грузам
    cargo_stats = pd.DataFrame({"Количество": cargo_counts})
    cargo_table = ax_table.table(
        cellText=[["Группа груза", "Количество"]] + cargo_stats.reset_index().values.tolist(),
        colWidths=[0.4, 0.2],
        cellLoc='center',
        bbox=[0.1, 0.55, 0.8, 0.4]
    )

    # Таблица по исполнителям
    driver_stats = pd.DataFrame({"Количество": driver_counts})
    driver_table = ax_table.table(
        cellText=[["ФИО исполнителя", "Количество"]] + driver_stats.reset_index().values.tolist(),
        colWidths=[0.4, 0.2],
        cellLoc='center',
        bbox=[0.1, 0.05, 0.8, 0.4]
    )

    # Стилизация таблиц
    for table in [cargo_table, driver_table]:
        table.auto_set_font_size(False)
        table.set_fontsize(12)
        for (row, col), cell in table.get_celld().items():
            if row == 0:
                cell.set_facecolor('#f0f0f0')
                cell.set_text_props(weight='bold')

def get_workshop_orders(orders, location_attr):
    """Фильтрует заказы для цехов"""
    return [
        o for o in orders
        if getattr(o, location_attr) and
        getattr(o, location_attr).department and
        getattr(o, location_attr).department.departmentType and
        getattr(o, location_attr).department.departmentType.department_type_id == 1
        and o.cargoType
    ]

def create_driver_orders_diagram(period_str: str, driver_stats: dict) -> BytesIO:
    """Диаграмма по водителям (только количество заказов по типам)"""
    fig, ax1 = create_figure_with_subplots(1, (17, 15))
    fig.suptitle(period_str, fontsize=16, fontweight='bold')
    # График распределения заказов по типам грузов
    stats = driver_stats
    stats['driver_cargo'].plot(
        kind='bar',
        stacked=True,
        ax=ax1,
        color=plt.cm.Set3(range(len(stats['driver_cargo'].columns))),
        edgecolor='black'
    )
    ax1.set_title("Количество заказов по водителям с разбиением по типам грузов")
    ax1.set_xlabel("Водитель")
    ax1.set_ylabel("Количество заказов")
    ax1.tick_params(axis='x', rotation=45)
    ax1.grid(True, axis='y')
    ax1.legend(title="Группа груза", bbox_to_anchor=(1.05, 1), loc='upper left')
    for i, driver in enumerate(stats['driver_cargo'].index):
        cumulative_height = 0
        for j, cargo_type in enumerate(stats['driver_cargo'].columns):
            value = stats['driver_cargo'].loc[driver, cargo_type]
            if value > 0:
                cumulative_height += value
                text_y = cumulative_height - (value / 2)
                ax1.text(i, text_y, int(value), ha='center', va='center', fontsize=8, color='black')
        total = stats['driver_cargo'].sum(axis=1)[driver]
        ax1.text(i, total + 0.5, int(total), ha='center', va='bottom')
    return save_figure_to_buffer(fig)

def create_driver_time_diagram(period_str: str, driver_stats: dict) -> BytesIO:
    """Диаграмма по среднему времени выполнения заказов по водителям"""
    fig, ax2 = create_figure_with_subplots(1, (17, 13))
    fig.suptitle(period_str, fontsize=16, fontweight='bold')
    stats = driver_stats
    bars = ax2.bar(
        stats['driver_time'].index,
        stats['driver_time'].values / 60,  # Переводим секунды в минуты
        color='lightgreen',
        edgecolor='black'
    )
    ax2.set_title("Среднее время выполнения заказов по водителям (в минутах)")
    ax2.set_xlabel("Водитель")
    ax2.set_ylabel("Среднее время (мин)")
    ax2.tick_params(axis='x', rotation=45)
    ax2.grid(True, axis='y')
    for bar in bars:
        yval = bar.get_height()
        ax2.text(bar.get_x() + bar.get_width()/2, yval + 0.5, f"{yval:.1f}", ha='center', va='bottom')
    return save_figure_to_buffer(fig)

def create_driver_diagram(period_str: str, driver_stats: dict) -> BytesIO:
    """Создает диаграмму по водителям"""
    fig, (ax1, ax2) = create_figure_with_subplots(2, (17, 19), [1, 1])
    fig.suptitle(period_str, fontsize=16, fontweight='bold')
    plot_driver_stats(ax1, ax2, driver_stats)
    return save_figure_to_buffer(fig)

def create_departments_with_buildings_diagram(period_str: str, from_stats: pd.DataFrame, to_stats: pd.DataFrame) -> BytesIO:
    """Создает диаграмму по цехам с корпусами"""
    fig, (ax1, ax2) = create_figure_with_subplots(2, (35, 19), [1, 1])
    fig.suptitle(period_str, fontsize=16, fontweight='bold')
    plot_grouped_bars(ax1, from_stats, "Поступление заказов из цехов по типам (c корпусами)", "Цех (откуда)", "tab20", True)
    plot_grouped_bars(ax2, to_stats, "Поступление заказов в цеха по типам (c корпусами)", "Цех (куда)", "tab20c", True)
    return save_figure_to_buffer(fig)

def create_departments_diagram(period_str: str, from_stats: pd.DataFrame, to_stats: pd.DataFrame) -> BytesIO:
    """Создает диаграмму по цехам без корпусов"""
    fig, (ax1, ax2) = create_figure_with_subplots(2, (35, 19), [1, 1])
    fig.suptitle(period_str, fontsize=16, fontweight='bold')
    plot_grouped_bars(ax1, from_stats, "Поступление заказов из цехов по типам", "Цех (откуда)", "tab20", False)
    plot_grouped_bars(ax2, to_stats, "Поступление заказов в цеха по типам", "Цех (куда)", "tab20c", False)
    return save_figure_to_buffer(fig)

def create_legend_diagram(cargo_counts: pd.Series, driver_counts: pd.Series) -> BytesIO:
    """Создает диаграмму с легендой"""
    fig = plt.figure(figsize=(12, 8))
    create_legend_table(fig, cargo_counts, driver_counts)
    return save_figure_to_buffer(fig)

def create_hierarchical_stats(df: pd.DataFrame, is_buildings: bool) -> pd.DataFrame:
    """Создает иерархическую статистику"""
    if df.empty:
        return pd.DataFrame()
    
    if is_buildings:
        # Группируем по корпусам, цехам и типам грузов
        grouped = df.groupby(['Корпус', 'Цех', 'Тип груза']).size().unstack(fill_value=0)
        return grouped.sort_index(level=[0, 1])
    else:
        # Группируем только по цехам и типам грузов
        grouped = df.groupby(['Цех', 'Тип груза']).size().unstack(fill_value=0)
        return grouped.sort_index()

def plot_grouped_bars(ax, stats_df: pd.DataFrame, title: str, xlabel: str, colormap: str, is_buildings: bool):
    """Строит сгруппированные столбчатые диаграммы"""
    if stats_df.empty:
        ax.text(0.5, 0.5, 'Нет данных', ha='center', va='center')
        ax.set_title(title)
        return
    # Подготовка позиций
    positions = []
    xtick_labels = []
    corpus_labels = []
    current_pos = 0
    if is_buildings:
        for corpus_name, corpus_group in stats_df.groupby(level=0):
            n_shops = len(corpus_group)
            shop_positions = range(current_pos, current_pos + n_shops)
            positions.extend(shop_positions)
            xtick_labels.extend(corpus_group.index.get_level_values(1).tolist())
            corpus_center = (shop_positions[0] + shop_positions[-1]) / 2
            corpus_labels.append((corpus_center, corpus_name))
            current_pos += n_shops + 1
    else:
        positions = range(len(stats_df))
        xtick_labels = stats_df.index.get_level_values(0).tolist()
    bottom = np.zeros(len(positions))
    colors = plt.get_cmap(colormap, len(stats_df.columns)).colors
    for i, col in enumerate(stats_df.columns):
        values = []
        if is_buildings:
            for corpus_name, corpus_group in stats_df.groupby(level=0):
                values.extend(corpus_group[col].values)
        else:
            values = stats_df[col].values
        bars = ax.bar(positions, values, bottom=bottom, 
               color=colors[i], 
               edgecolor='black',
               label=col)
        # Подписи внутри сегментов только для текущего слоя
        for idx, (pos, value) in enumerate(zip(positions, values)):
            if value > 0:
                y = bottom[idx] + value / 2
                ax.text(pos, y, str(int(value)), ha='center', va='center', fontsize=8, color='black')
        bottom += np.array(values)
    ax.set_title(title)
    ax.set_xlabel(xlabel)
    ax.set_ylabel("Количество")
    ax.set_xticks(positions)
    ax.set_xticklabels(xtick_labels, rotation=45, ha="right")
    ax.grid(True, axis='y')
    ax.legend(title="Тип груза", bbox_to_anchor=(1.05, 1), loc='upper left')
    if is_buildings:
        for center, corpus_name in corpus_labels:
            ax.text(center, -0.1 * ax.get_ylim()[1], corpus_name, 
                    ha='center', va='top',
                    fontsize=10, fontweight='bold', color='darkblue')
    # Подписи total только один раз (сумма по столбцу)
    for idx, pos in enumerate(positions):
        total = bottom[idx]
        if total > 0:
            ax.text(pos, total + 0.5, str(int(total)), 
                    ha='center', va='bottom',
                    fontsize=9, fontweight='bold')

def create_driver_diagram_weighted(period_str: str, driver_stats: dict) -> BytesIO:
    try:
        fig, ax1 = create_figure_with_subplots(1, (17, 15))
        fig.suptitle(period_str + ' (с учетом веса)', fontsize=16, fontweight='bold')
        stats = driver_stats
        stats['driver_cargo'].plot(
            kind='bar',
            stacked=True,
            ax=ax1,
            color=plt.cm.Set3(range(len(stats['driver_cargo'].columns))),
            edgecolor='black'
        )
        ax1.set_title("Количество заказов по водителям с учетом веса (с разбиением по типам)")
        ax1.set_xlabel("Водитель")
        ax1.set_ylabel("Суммарный вес")
        ax1.tick_params(axis='x', rotation=45)
        ax1.grid(True, axis='y')
        ax1.legend(title="Группа груза", bbox_to_anchor=(1.05, 1), loc='upper left')
        for i, driver in enumerate(stats['driver_cargo'].index):
            cumulative_height = 0
            for j, cargo_type in enumerate(stats['driver_cargo'].columns):
                value = stats['driver_cargo'].loc[driver, cargo_type]
                if value > 0:
                    cumulative_height += value
                    text_y = cumulative_height - (value / 2)
                    ax1.text(i, text_y, f"{value:.1f}", ha='center', va='center', fontsize=8, color='black')
            total = stats['driver_cargo'].sum(axis=1)[driver]
            ax1.text(i, total + 0.5, f"{total:.1f}", ha='center', va='bottom')
        return save_figure_to_buffer(fig)
    except Exception as e:
        logging.error(f"[create_driver_diagram_weighted] Ошибка при построении диаграммы: {e}")
        raise

async def create_driver_stats_weighted_for_export(orders, session):
    logging.info(f"[create_driver_stats_weighted_for_export] Start. Orders count: {len(orders)}")
    from collections import defaultdict
    driver_cargo_weighted = defaultdict(lambda: defaultdict(float))
    driver_counts_weighted = defaultdict(float)
    cargo_counts_weighted = defaultdict(float)

    async def get_weighted(order):
        try:
            if order.executor and order.cargoType and getattr(order.executor, 'is_denied', True) == False and getattr(order.executor, 'role_id', None) == 2:
                #coeff = await get_weight_coefficient_by_order_id(order.id_order) - старый расчет коефицента
                weight = getattr(order, 'cargo_weight', 1)
                logging.debug(f"Order {order.id_order}: fio={order.executor.fio}, cargo={order.cargoType.cargo_type_name}, weight={weight}")
                return (order.executor.fio, order.cargoType.cargo_type_name, weight)
        except Exception as e:
            logging.error(f"Ошибка при вычислении коэффициента веса для заказа {getattr(order, 'id_order', None)}: {e}")
        return None

    tasks = [get_weighted(order) for order in orders]
    results = await asyncio.gather(*tasks)
    for res in results:
        if res:
            fio, cargo, weight = res
            driver_cargo_weighted[fio][cargo] += weight
            driver_counts_weighted[fio] += weight
            cargo_counts_weighted[cargo] += weight

    import pandas as pd
    df_driver_cargo = pd.DataFrame(driver_cargo_weighted).T.fillna(0)
    df_driver_cargo = df_driver_cargo[df_driver_cargo.sum(axis=1) > 0]
    logging.info(f"[create_driver_stats_weighted_for_export] Done. Drivers: {list(driver_counts_weighted.keys())}")
    return {
        'driver_cargo': df_driver_cargo,
        'driver_counts': pd.Series(driver_counts_weighted),
        'cargo_counts': pd.Series(cargo_counts_weighted)
    }

@connection
async def export_diagrama(session,
    diogramType,
    date_from: datetime = None,
    date_to: datetime = datetime.today() + timedelta(days = 1)) -> BufferedInputFile:
    logging.info(f"[export_diagrama] Начало экспорта диаграммы. Тип: {diogramType}, с {date_from} по {date_to}")
    try:
        logging.info("[export_diagrama] Получение заказов из базы данных")
        stmt = (
            select(tb.Order)
            .options(
                joinedload(tb.Order.cargoType),
                joinedload(tb.Order.executor),
                joinedload(tb.Order.depart_loc_ref).joinedload(tb.DepartmentBuilding.department).joinedload(tb.Department.departmentType),
                joinedload(tb.Order.depart_loc_ref).joinedload(tb.DepartmentBuilding.building),
                joinedload(tb.Order.goal_loc_ref).joinedload(tb.DepartmentBuilding.department).joinedload(tb.Department.departmentType),
                joinedload(tb.Order.goal_loc_ref).joinedload(tb.DepartmentBuilding.building)
            )
            .where(and_(
                tb.Order.order_status_id == 3,
                tb.Order.completion_time.isnot(None),
                tb.Order.pickup_time.isnot(None),
                tb.Order.completion_time >= date_from,
                tb.Order.completion_time <= date_to
            ))
        )
        result = await session.execute(stmt)
        orders = result.scalars().all()
        logging.info(f"[export_diagrama] Получено заказов: {len(orders)}")
        if not orders:
            logging.error("[export_diagrama] Нет выполненных заказов за указанный период")
            raise ValueError("Нет выполненных заказов за указанный период")

        logging.info("[export_diagrama] Формирование driver_stats")
        driver_stats = create_driver_stats(orders)
        logging.info("[export_diagrama] Формирование driver_stats_weighted")
        driver_stats_weighted = await create_driver_stats_weighted_for_export(orders, session)

        logging.info("[export_diagrama] Формирование orders_from_workshops и orders_to_workshops")
        orders_from_workshops = get_workshop_orders(orders, 'depart_loc_ref')
        orders_to_workshops = get_workshop_orders(orders, 'goal_loc_ref')
        from_df = get_order_data(orders_from_workshops, 'depart_loc_ref')
        to_df = get_order_data(orders_to_workshops, 'goal_loc_ref')
        from_stats = create_hierarchical_stats(from_df, True)
        to_stats = create_hierarchical_stats(to_df, True)
        from_stats_short = create_hierarchical_stats(from_df, False)
        to_stats_short = create_hierarchical_stats(to_df, False)

        period_str = f"Период: {date_from.strftime('%d.%m.%Y')} — {date_to.strftime('%d.%m.%Y')}"

        diagrams = {
            'drivers': [
                create_driver_orders_diagram(period_str, driver_stats)
            ],
            'time': [
                create_driver_time_diagram(period_str, driver_stats)
            ],
            'driversWithWeight': [
                create_driver_diagram_weighted(period_str, driver_stats_weighted)
            ],
            'depBuild': [
                create_departments_with_buildings_diagram(period_str, from_stats, to_stats),
                create_legend_diagram(driver_stats['cargo_counts'], driver_stats['driver_counts'])
            ],
            'dep': [
                create_departments_diagram(period_str, from_stats_short, to_stats_short),
                create_legend_diagram(driver_stats['cargo_counts'], driver_stats['driver_counts'])
            ],
            'all': [
                create_driver_orders_diagram(period_str, driver_stats),
                create_driver_time_diagram(period_str, driver_stats),
                create_driver_diagram_weighted(period_str, driver_stats_weighted),
                create_departments_with_buildings_diagram(period_str, from_stats, to_stats),
                create_departments_diagram(period_str, from_stats_short, to_stats_short),
                create_legend_diagram(driver_stats['cargo_counts'], driver_stats['driver_counts'])
            ]
        }

        if diogramType not in diagrams:
            logging.error(f"[export_diagrama] Неизвестный тип диаграммы: {diogramType}")
            raise ValueError(f"Неизвестный тип диаграммы: {diogramType}")

        logging.info(f"[export_diagrama] Формирование файлов для типа: {diogramType}")
        result = []
        for i, buffer in enumerate(diagrams[diogramType]):
            if i == 1 and diogramType in ['depBuild', 'dep']:
                filename = "Легенда к диаграмме цехов"
            else:
                filename = f"{diogramType}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png"
            logging.info(f"[export_diagrama] Добавление файла: {filename}")
            result.append(BufferedInputFile(buffer.getvalue(), filename=filename))
        logging.info(f"[export_diagrama] Успешно завершено. Файлов: {len(result)}")
        return result
    except Exception as e:
        logging.error(f"[export_diagrama] Ошибка при экспорте диаграммы: {e}")
        raise

# Получаем данные с учетом корпусов
def get_order_data(orders, location_attr):
    data = []
    for order in orders:
        loc = getattr(order, location_attr)
        if loc:  # Если есть привязка к месту
            building_name = loc.building.building_name if loc.building else "Без корпуса"
            department_name = loc.department.department_name if loc.department else "Без цеха"
            cargo_type = order.cargoType.cargo_type_name if order.cargoType else "Без типа"
            
            data.append({
                'Корпус': building_name,
                'Цех': department_name,
                'Тип груза': cargo_type,
                'Количество': 1
            })
    return pd.DataFrame(data)

@connection
async def get_user_id(session: AsyncSession, tg_id: int) -> int:
    if tg_id in user_cache:  
        return user_cache[tg_id]  
    
    user = await session.scalar(select(tb.User).where(tb.User.tgId == tg_id))
    user_cache[tg_id] = user.id_user
    return user.id_user

@connection
async def save_location(session: AsyncSession, user_id, latitude, longitude, timestamp):

    new_loc = tb.UserLocation(
        user_id=user_id,
        latitude=latitude,
        longitude=longitude,
        timestamp=timestamp,
    )

    session.add(new_loc)

@connection
async def get_map(session: AsyncSession, tg_id, date):
    
    #user_id = await get_user_id(tg_id = tg_id)
    user_id = 2
    date = datetime.today() - timedelta(days=1)

    stmt = (
        select(tb.UserLocation)
        .where(and_ (func.date(tb.UserLocation.timestamp) == date.date(),
                     tb.UserLocation.user_id == user_id))
        .order_by((tb.UserLocation.created_at))
    )

    results = await session.stream(stmt)
    locations = await results.scalars().all()

    if len(locations) < 2:
        return "Недостаточно данных для построения маршрута. Требуется хотя бы две точки."
    
    coordinates = [(loc.longitude, loc.latitude) for loc in locations]
    coordinates = coordinates[::10]
    route_points = ','.join([f'{lat},{lon}' for lat, lon in coordinates])

    loop = asyncio.get_event_loop()
    image_data = await loop.run_in_executor(None, generate_map, coordinates)

    return BufferedInputFile(file=image_data, filename="route_map.png")

def generate_map(coordinates):
    
    
    # Создаём объект карты с размерами 800x600 пикселей
    m = StaticMap(800, 600)
    
    # Создаём линию маршрута: синяя, толщиной 5 пикселей
    line = Line(coordinates, 'blue', 5)
    m.add_line(line)

    # Рендерим карту в изображение
    image = m.render()

    # Сохраняем изображение в буфер памяти
    image_io = BytesIO()
    image.save(image_io, 'PNG')
    image_io.seek(0)

    return image_io.getvalue()

@connection
async def change_role(session: AsyncSession, data, id_role):
        
        user = await session.scalar(select(tb.User).where(tb.User.tgId == data["tg_id"]))
        new_data={
            "role_id": id_role
        }

        stmt = (
            update(tb.User)
            .where(tb.User.id_user == user.id_user)
            .values(**new_data)
        )

        await session.execute(stmt)

@connection
async def set_driver_rate(session: AsyncSession, orderId, rate):
    updates = {
        "driver_rate": rate
    }

    stmt = (
        update(tb.Order)
        .where(tb.Order.id_order == orderId)
        .values(**updates)
    )

    await session.execute(stmt)


def get_dep_build_id(dep_id: int, build_id: int) -> int:
    for entry in dep_build_cache["department_buildings"]:
        if entry["department_id"] == dep_id and entry["building_id"] == build_id:
            return entry["id"] 
    raise ValueError(f"Не найдена запись DepartmentBuilding в кеше. Атрубуты: department_id={dep_id} and building_id={build_id}")

def get_dep_build_description(dep_build_id: int) -> str:
    for entry in dep_build_cache["department_buildings"]:
        if entry["id"] == dep_build_id:
            return entry["description"] 
    raise ValueError(f"Не найдено описание для dep_build_id = {dep_build_id}")

def get_dep_name(dep_id: int, isWithTypeName: False) -> str:
    deps = dep_build_cache.get("departments", [])
    res = ''
    for dep in deps:
        if dep["idDepartment"] == dep_id:
            if isWithTypeName:
                match dep["typeId"]:
                    case 1: res = "Цех "
                    case 2: res = "Отдел "
                    case _: res = "Тип не определен. Номер: "
            return res + dep["departmentName"]
    raise ValueError(f"Не найдена запись Department в кеше. Атрибут: id={dep_id}")

def get_build_name(build_id: int) -> str:
    bulds = dep_build_cache.get("buildings", [])
    for buld in bulds:
        if buld["idBuilding"] == build_id:
            return buld["buildingName"]
    raise ValueError(f"Не найдена запись Building в кеше. Атрибут: id={build_id}")

def get_bilds_List(dep_id: int) -> List[dict]:
    return [{"id": entry["id"], "building_id": entry["building_id"]}
            for entry in dep_build_cache["department_buildings"]
            if entry["department_id"] == dep_id]

def get_bilds_List_all() -> List[dict]:
    return [{"id": entry["id"], "building_id": entry["building_id"]}
            for entry in dep_build_cache["department_buildings"]]

def get_dep_List(dep_type_id: int) -> List[dict]:
    departments = dep_build_cache.get("departments", [])
    return [{'id': dep["idDepartment"], 'name': dep["departmentName"]}
            for dep in departments if dep["typeId"] == dep_type_id]

def get_dep_build_input(dep_build_id: int) -> str:
    for entry in dep_build_cache["department_buildings"]:
        if entry["id"] == dep_build_id:
            dep_name = get_dep_name(entry["department_id"], isWithTypeName=True)
            build_name = get_build_name(entry["building_id"])
            res = f'{dep_name}, корпус {build_name}, {entry["description"]}'
            return res

@connection
async def dep_build_set(session: AsyncSession):
    deps = list(await session.scalars(select(tb.Department).order_by(tb.Department.department_name)))
    build = list(await session.scalars(select(tb.Building).order_by(tb.Building.building_name)))
    dep_build = list(await session.scalars(select(tb.DepartmentBuilding)))

    dep_build_cache.clear()
    dep_build_cache["department_buildings"] = [
        {"id": db.department_building_id, "department_id": db.department_id, "building_id": db.building_id, "description": db.description}
        for db in dep_build
    ]
    dep_build_cache["departments"] = [
        {"idDepartment": dep.department_id, "departmentName": dep.department_name, "typeId": dep.department_type_id}
        for dep in deps
    ]
    dep_build_cache["buildings"] = [
        {"idBuilding": b.building_id, "buildingName": b.building_name}
        for b in build
    ]

@connection
async def get_cargo_type_list(session: AsyncSession):
    rows = await session.scalars(select(tb.CargoType).order_by(tb.CargoType.id_cargo_type))
    return [{"id": row.id_cargo_type, "label": row.cargo_type_name, "coefficent": row.ratio} for row in rows]



@connection
async def get_time_coeffs(session: AsyncSession):
    rows = await session.scalars(select(tb.TimeCoeff).order_by(tb.TimeCoeff.value))
    return [{"id": row.time_coefficent_id, "label": f"{row.value} мин", "coefficent": row.coefficent} for row in rows]

@connection
async def update_time_coeff(session: AsyncSession, id: int, coeff: float):
    stmt = update(tb.TimeCoeff).where(tb.TimeCoeff.time_coefficent_id == id).values(coefficent=coeff)
    await session.execute(stmt)

@connection
async def get_weight_coeffs(session: AsyncSession):
    rows = await session.scalars(select(tb.WeightCoeff).order_by(tb.WeightCoeff.value))
    return [{"id": row.weight_coefficent_id, "label": f"{row.value} кг", "coefficent": row.coefficent} for row in rows]

@connection
async def update_weight_coeff(session: AsyncSession, id: int, coeff: float):
    stmt = update(tb.WeightCoeff).where(tb.WeightCoeff.weight_coefficent_id == id).values(coefficent=coeff)
    await session.execute(stmt)

@connection
async def update_ratio(session: AsyncSession, id, ratio):
    updates ={
        "ratio": ratio,
    }

    stmt = (
        update(tb.CargoType)
        .where(tb.CargoType.id_cargo_type == int(id))
        .values(**updates)
    )

    await session.execute(stmt)

@connection
async def add_ratio(session: AsyncSession, coeff_type, value):
    match coeff_type:
        case "cargo":
            data_save = tb.CargoType(
                cargo_type_name = value,
                ratio = 1.0,
            )

        case "time":
            data_save = tb.TimeCoeff(
                value = value,
                coefficent = 1.0,
            )

        case "weight":
            data_save = tb.WeightCoeff(
                value = value,
                coefficent = 1.0,
            )
            
    session.add(data_save)

@connection
async def get_stuff_List_mes(session: AsyncSession, roleId: int):
    stmt = (
        select(tb.User)
        .where(tb.User.role_id == roleId)
        .order_by(tb.User.fio)
    )

    res = await session.execute(stmt)
    stuffList = res.scalars().all()

    mes = "\n"
    for stuff in stuffList:
        mes = mes + stuff.fio
        if roleId == 2:
            rate, count = await get_driver_rate(stuff.id_user)
            mes = mes + f'\n\t⭐ рейтинг: {str(rate)} \n\t📦 всего заказов с оценкой: {str(count)}'
        mes = mes +"\n\n"
    return mes

@connection
async def get_driver_rate(session: AsyncSession, driverId: int):
    stmt = select(
        func.avg(tb.Order.driver_rate),
        func.count(tb.Order.driver_rate)
        ).where(
        tb.Order.driver_id == driverId,
        tb.Order.driver_rate.isnot(None)
    )

    result = await session.execute(stmt)
    average_rate, count = result.one_or_none()

    average_rate = round(average_rate, 2) if average_rate is not None else 0.0
    count = count or 0
    return average_rate, count

@connection
async def get_admins_for_alarm(session: AsyncSession):
    stmt = (
        select(tb.User)
        .where(tb.User.role_id == 4)
    )

    result = await session.execute(stmt)
    admins = result.scalars().all()
    admin_ids = [admin.tgId for admin in admins]
    return admin_ids

@connection
async def get_drivers_payment(session: AsyncSession, last_month_12 = None, current_month_12 = None):
    today = datetime.now()

    total_bonus = 200000
    salary = 50000

    if (current_month_12 is None):
        current_month_12 = today.replace(day=12)
        if today.month == 1:  # Если январь, берём декабрь прошлого года
            last_month_12 = today.replace(year=today.year - 1, month=12, day=12)
        else:
            last_month_12 = today.replace(month=today.month - 1, day=12)
    else:
        last_month_12 = datetime.strptime(last_month_12, '%d.%m.%Y')
        current_month_12 = datetime.strptime(current_month_12, '%d.%m.%Y')

    drivers = await session.execute(
            select(tb.User.id_user, tb.User.fio)
            .where(tb.User.role_id == 2))
    
    drivers = drivers.all()

    drivers_names = {id_user: fio for id_user, fio in drivers}
    drivers_dict = {id_user: 0 for id_user, _ in drivers}

    stmt = (
        select(tb.Order)
        .options(joinedload(tb.Order.executor))
        .options(joinedload(tb.Order.cargoType))
        .where(
            tb.Order.completion_time >= last_month_12,
            tb.Order.completion_time < current_month_12
        )
    )
    result = await session.execute(stmt)
    orders = result.scalars().all()

    for order in orders:
        # Учитываем только если исполнитель существует и его роль == 2
        if order.executor and getattr(order.executor, 'role_id', None) == 2:
            drivers_dict[order.executor.id_user] = drivers_dict[order.executor.id_user] + 1 * order.cargoType.ratio

    total_orders = sum(drivers_dict.values())
    if total_orders == 0:
        return "Нет заказов за период."
    sum_on_order = total_bonus / total_orders
    
    mes = f'📊Расчёт зарплат за период с {last_month_12.strftime("%d-%m-%Y")} по {current_month_12.strftime("%d-%m-%Y")}\n📦Всего отвезено заказов (с учетом коэффициентов): {total_orders}\n💰Общая сумма на премии: {total_bonus}\n📈Сумма премии на единицу заказа: {sum_on_order:.2f}\n💵Фиксированный оклад: {salary}\n\n'    
    for driver_id in drivers_dict:
        bonus = drivers_dict[driver_id] * sum_on_order
        total_salary = bonus + salary
        mes += f'👤{drivers_names[driver_id]}\n💵Премия: {bonus:.2f}\n💼Зарплата: {total_salary:.2f} \n\n'

    return mes
    

@connection
async def add_department(session: AsyncSession, name: str, type_id: int):
    dep = tb.Department(department_name=name, department_type_id=type_id)
    session.add(dep)
    await session.flush()
    return dep.department_id

@connection
async def add_building(session: AsyncSession, name: str):
    existing = await session.scalar(select(tb.Building).where(tb.Building.building_name == name))
    if existing:
        return existing.building_id
    build = tb.Building(building_name=name)
    session.add(build)
    await session.flush()
    return build.building_id

@connection
async def add_department_building(session: AsyncSession, department_id: int, building_id: int, description: str = None):
    dep_build = tb.DepartmentBuilding(department_id=department_id, building_id=building_id, description=description)
    session.add(dep_build)
    await session.flush()
    return dep_build.department_building_id

@connection
async def add_department_and_building(session: AsyncSession, department_name: str, department_type_id: int, building_name: str, description: str = None):
    dep = tb.Department(department_name=department_name, department_type_id=department_type_id)
    session.add(dep)
    await session.flush()
    build = tb.Building(building_name=building_name)
    session.add(build)
    await session.flush()
    dep_build = tb.DepartmentBuilding(department_id=dep.department_id, building_id=build.building_id, description=description)
    session.add(dep_build)
    await session.flush()
    return dep.department_id, build.building_id, dep_build.department_building_id

@connection
async def get_weight_coefficient_by_order_id(session: AsyncSession, order_id: int) -> float:
    logging.info(f"[get_weight_coefficient_by_order_id] Получение коэффициента веса для заказа {order_id}")
    try:
        order = await session.scalar(select(tb.Order).where(tb.Order.id_order == order_id))
        if not order:
            raise ValueError(f"Order with id {order_id} not found")
        weight = order.cargo_weight
        coeff = await session.scalar(select(tb.WeightCoeff))
        num = weight // coeff.value
        result = num * coeff.coefficent
        logging.debug(f"[get_weight_coefficient_by_order_id] Вес: {weight}, шаг: {coeff.value}, коэффициент: {coeff.coefficent}, результат: {result}")
        return result
    except Exception as e:
        logging.error(f"[get_weight_coefficient_by_order_id] Ошибка: {e}")
        raise

@connection
async def export_delivery_points_to_excel(session: AsyncSession) -> BufferedInputFile:
    """
    Экспорт списка точек доставки в Excel файл
    """
    try:
        # Формируем запрос с загрузкой связанных данных
        stmt = (
            select(tb.DepartmentBuilding)
            .options(joinedload(tb.DepartmentBuilding.department))
            .options(joinedload(tb.DepartmentBuilding.building))
            .order_by(tb.DepartmentBuilding.department_id, tb.DepartmentBuilding.building_id)
        )

        result = await session.execute(stmt)
        delivery_points = result.scalars().all()

        if not delivery_points:
            raise ValueError("В базе данных нет точек доставки")

        # Подготавливаем данные для DataFrame
        data = []
        for point in delivery_points:
            data.append({
                'Id': point.department_building_id,
                'Отдел': point.department.department_name if point.department else "Не указан",
                'Здание': point.building.building_name if point.building else "Не указано",
                'Описание': point.description if point.description else "Не указано"
            })

        df = pd.DataFrame(data)

        # Создаем Excel файл
        excel_file = BytesIO()
        with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Delivery Points')
            worksheet = writer.sheets['Delivery Points']

            # Настраиваем ширину колонок
            for idx, column in enumerate(df.columns, 1):
                max_length = max(df[column].astype(str).map(len).max(), len(column)) + 2
                worksheet.column_dimensions[get_column_letter(idx)].width = max_length

        excel_file.seek(0)
        
        filename = f"delivery_points_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        file_for_telegram = BufferedInputFile(
            file=excel_file.getvalue(),
            filename=filename
        )

        return file_for_telegram

    except Exception as e:
        logging.error(f"Ошибка при экспорте точек доставки в Excel: {e}")
        raise

